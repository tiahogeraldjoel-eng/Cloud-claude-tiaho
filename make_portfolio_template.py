"""
Générateur du fichier Excel de suivi de portefeuille BRVM.
Usage : python make_portfolio_template.py
Produit : BRVM_Portefeuille.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ─── Couleurs ────────────────────────────────────────────────────────────────

C_HEADER   = "1F3864"   # bleu nuit
C_SIGNAL_H = "C00000"   # rouge HIGH
C_SIGNAL_M = "E36C09"   # orange MEDIUM
C_SIGNAL_L = "FFD966"   # jaune LOW
C_GREEN    = "70AD47"   # vert
C_RED      = "FF0000"
C_ORANGE   = "FF7F00"
C_LIGHT    = "D9E1F2"   # bleu clair ligne paire
C_WHITE    = "FFFFFF"
C_GOLD     = "FFD700"

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, value, bg=C_HEADER, fg="FFFFFF", bold=True, wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill    = PatternFill("solid", fgColor=bg)
    c.font    = Font(color=fg, bold=bold, name="Calibri", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border  = border()
    return c

def data_cell(ws, row, col, value=None, bg=C_WHITE, bold=False, num_fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill    = PatternFill("solid", fgColor=bg)
    c.font    = Font(name="Calibri", size=10, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border  = border()
    if num_fmt:
        c.number_format = num_fmt
    return c

# ─── Feuille 1 : Positions ouvertes ─────────────────────────────────────────

def make_positions(wb):
    ws = wb.create_sheet("📊 Positions", 0)
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:O1")
    t = ws["A1"]
    t.value = "BRVM — Suivi des Positions Ouvertes"
    t.fill  = PatternFill("solid", fgColor=C_HEADER)
    t.font  = Font(color="FFFFFF", bold=True, name="Calibri", size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sous-titre
    ws.merge_cells("A2:O2")
    s = ws["A2"]
    s.value = "Mise à jour : saisir les prix actuels manuellement — ou configurer Power Query (onglet Cours_Live)"
    s.fill  = PatternFill("solid", fgColor="2E75B6")
    s.font  = Font(color="FFFFFF", italic=True, name="Calibri", size=9)
    s.alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes
    headers = [
        ("Date\nentrée",    10),  # A
        ("Ticker",           8),  # B
        ("Nom société",     22),  # C
        ("Signal\n(H/M/L)", 9),   # D
        ("Qtité\nachetée",  9),   # E
        ("Prix\nentrée",    10),  # F
        ("Coût\ntotal",     11),  # G
        ("Stop-loss\n−3%",  10),  # H
        ("Objectif\n+4%",   10),  # I
        ("Max BRVM\n+7.5%", 10),  # J
        ("Prix\nactuel",    10),  # K  ← saisir ici
        ("P&L\nFCFA",       11),  # L
        ("P&L\n%",           8),  # M
        ("Statut",          16),  # N
        ("Notes",           20),  # O
    ]
    ws.row_dimensions[3].height = 32
    for col, (hdr, w) in enumerate(headers, start=1):
        hdr_cell(ws, 3, col, hdr, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    # 15 lignes de données avec formules
    for r in range(4, 19):
        bg = C_LIGHT if r % 2 == 0 else C_WHITE
        row = str(r)

        # A : date
        c = data_cell(ws, r, 1, bg=bg, align="center")
        c.number_format = "DD/MM/YYYY"

        # B : ticker
        data_cell(ws, r, 2, bg=bg, bold=True, align="center")

        # C : nom (lookup dans Référentiel si dispo, sinon saisie libre)
        data_cell(ws, r, 3, bg=bg, align="left")

        # D : signal
        data_cell(ws, r, 4, bg=bg, align="center")

        # E : quantité achetée
        data_cell(ws, r, 5, bg=bg, num_fmt="#,##0", align="center")

        # F : prix entrée
        c = data_cell(ws, r, 6, bg=bg, num_fmt="#,##0", align="right")

        # G : coût total = E × F
        c = data_cell(ws, r, 7,
            value=f"=IF(E{row}*F{row}=0,\"\",E{row}*F{row})",
            bg=bg, num_fmt="#,##0", align="right", bold=True)

        # H : stop-loss = F × 0.97
        data_cell(ws, r, 8,
            value=f"=IF(F{row}=0,\"\",ROUND(F{row}*0.97,0))",
            bg=bg, num_fmt="#,##0 \"F\"", align="right")

        # I : objectif = F × 1.04
        data_cell(ws, r, 9,
            value=f"=IF(F{row}=0,\"\",ROUND(F{row}*1.04,0))",
            bg=bg, num_fmt="#,##0 \"F\"", align="right")

        # J : max BRVM = F × 1.075
        data_cell(ws, r, 10,
            value=f"=IF(F{row}=0,\"\",ROUND(F{row}*1.075,0))",
            bg=bg, num_fmt="#,##0 \"F\"", align="right")

        # K : prix actuel — SAISIR ICI
        c = data_cell(ws, r, 11, bg="FFFACD", num_fmt="#,##0", align="right")
        c.fill = PatternFill("solid", fgColor="FFFACD")  # jaune pâle = zone de saisie

        # L : P&L FCFA = (K - F) × E
        data_cell(ws, r, 12,
            value=f"=IF(OR(E{row}=0,F{row}=0,K{row}=0),\"\",(K{row}-F{row})*E{row})",
            bg=bg, num_fmt="+#,##0;-#,##0;0", align="right", bold=True)

        # M : P&L % = (K - F) / F
        data_cell(ws, r, 13,
            value=f"=IF(OR(F{row}=0,K{row}=0),\"\",(K{row}-F{row})/F{row})",
            bg=bg, num_fmt="+0.00%;-0.00%;0.00%", align="center")

        # N : statut
        data_cell(ws, r, 14,
            value=(
                f'=IF(F{row}=0,"",IF(K{row}=0,"⏳ En attente",'
                f'IF(K{row}<=ROUND(F{row}*0.97,0),"🛑 STOP LOSS ATTEINT",'
                f'IF(K{row}>=ROUND(F{row}*1.075,0),"🚀 SORTIR (max BRVM)",'
                f'IF(K{row}>=ROUND(F{row}*1.04,0),"✅ SORTIR (objectif atteint)",'
                f'"⏳ TENIR")))))'
            ),
            bg=bg, align="center")

        # O : notes
        data_cell(ws, r, 15, bg=bg, align="left")

    # Mise en forme conditionnelle sur Statut (col N = 14)
    red_fill    = PatternFill("solid", fgColor="FFCCCC")
    green_fill  = PatternFill("solid", fgColor="CCFFCC")
    orange_fill = PatternFill("solid", fgColor="FFE0B2")
    rocket_fill = PatternFill("solid", fgColor="B3E5FC")

    rng = "N4:N18"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['ISNUMBER(SEARCH("STOP",N4))'], fill=red_fill,
        font=Font(bold=True, color=C_RED, name="Calibri", size=10)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['ISNUMBER(SEARCH("objectif",N4))'], fill=green_fill,
        font=Font(bold=True, color="006400", name="Calibri", size=10)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['ISNUMBER(SEARCH("max BRVM",N4))'], fill=rocket_fill,
        font=Font(bold=True, color="0D47A1", name="Calibri", size=10)))

    # P&L coloré
    ws.conditional_formatting.add("L4:L18", FormulaRule(
        formula=["L4>0"], fill=green_fill, font=Font(bold=True, color="006400", name="Calibri", size=10)))
    ws.conditional_formatting.add("L4:L18", FormulaRule(
        formula=["L4<0"], fill=red_fill, font=Font(bold=True, color=C_RED, name="Calibri", size=10)))

    # Légende signal col D
    ws.conditional_formatting.add("D4:D18", FormulaRule(
        formula=['D4="H"'], fill=PatternFill("solid", fgColor="FFCCCC")))
    ws.conditional_formatting.add("D4:D18", FormulaRule(
        formula=['D4="M"'], fill=PatternFill("solid", fgColor="FFE0B2")))
    ws.conditional_formatting.add("D4:D18", FormulaRule(
        formula=['D4="L"'], fill=PatternFill("solid", fgColor="FFFACD")))

    # Note d'utilisation
    ws.merge_cells("A20:O20")
    n = ws["A20"]
    n.value = (
        "MODE D'EMPLOI : 1) Remplis A-F quand tu passes un ordre après une alerte Telegram  "
        "2) Mets à jour K (Prix actuel) pendant la journée  "
        "3) Suis le Statut en col N  "
        "4) Transfère vers Historique quand tu clôtures"
    )
    n.fill = PatternFill("solid", fgColor="FFF3CD")
    n.font = Font(name="Calibri", size=9, italic=True, color="856404")
    n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[20].height = 28

    # Légende statuts
    legend = [
        (22, "Légende statuts :", C_HEADER, "FFFFFF", True),
        (23, "⏳ TENIR", "FFFFFF", "000000", False),
        (24, "✅ SORTIR (objectif atteint)", "CCFFCC", "006400", True),
        (25, "🚀 SORTIR (max BRVM)", "B3E5FC", "0D47A1", True),
        (26, "🛑 STOP LOSS ATTEINT → VENDRE IMMÉDIATEMENT", "FFCCCC", C_RED, True),
    ]
    for row_num, val, bg, fg, bold in legend:
        ws.merge_cells(f"A{row_num}:E{row_num}")
        c = ws[f"A{row_num}"]
        c.value = val
        c.fill  = PatternFill("solid", fgColor=bg)
        c.font  = Font(name="Calibri", size=10, bold=bold, color=fg)
        c.alignment = Alignment(vertical="center")
        c.border = border()

    ws.freeze_panes = "A4"
    return ws

# ─── Feuille 2 : Historique ───────────────────────────────────────────────────

def make_historique(wb):
    ws = wb.create_sheet("📁 Historique", 1)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    t = ws["A1"]
    t.value = "BRVM — Historique des Trades Clôturés"
    t.fill  = PatternFill("solid", fgColor="2E4057")
    t.font  = Font(color="FFFFFF", bold=True, name="Calibri", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        ("Date\nentrée", 12), ("Date\nsortie", 12), ("Ticker", 8),
        ("Nom", 22), ("Signal", 9), ("Qtité", 8),
        ("Prix\nentrée", 10), ("Prix\nsortie", 10), ("Coût", 11),
        ("Produit\nvente", 11), ("P&L\nFCFA", 11), ("P&L\n%", 8),
        ("Raison sortie", 18),
    ]
    ws.row_dimensions[2].height = 30
    for col, (hdr, w) in enumerate(headers, start=1):
        hdr_cell(ws, 2, col, hdr, bg="2E4057", wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    for r in range(3, 30):
        bg = C_LIGHT if r % 2 == 0 else C_WHITE
        row = str(r)
        for col in [1, 2]:
            c = data_cell(ws, r, col, bg=bg)
            c.number_format = "DD/MM/YYYY"
        for col in range(3, 14):
            data_cell(ws, r, col, bg=bg,
                num_fmt="#,##0" if col in [6,7,8,9,10] else
                        "+0.00%;-0.00%" if col == 12 else "General",
                align="right" if col in [6,7,8,9,10,11,12] else "center")

        # P&L FCFA = (sortie - entrée) × qtité
        data_cell(ws, r, 11,
            value=f"=IF(OR(G{row}=0,H{row}=0,F{row}=0),\"\",(H{row}-G{row})*F{row})",
            bg=bg, num_fmt="+#,##0;-#,##0;0", bold=True, align="right")
        # P&L %
        data_cell(ws, r, 12,
            value=f"=IF(OR(G{row}=0,H{row}=0),\"\",(H{row}-G{row})/G{row})",
            bg=bg, num_fmt="+0.00%;-0.00%", align="center")
        # Produit vente
        data_cell(ws, r, 10,
            value=f"=IF(OR(H{row}=0,F{row}=0),\"\",H{row}*F{row})",
            bg=bg, num_fmt="#,##0", align="right")

    # Total P&L
    ws.merge_cells("A31:J31")
    ws["A31"].value = "TOTAL P&L"
    ws["A31"].fill  = PatternFill("solid", fgColor="2E4057")
    ws["A31"].font  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    ws["A31"].alignment = Alignment(horizontal="right", vertical="center")
    c = ws["K31"]
    c.value  = "=SUM(K3:K30)"
    c.fill   = PatternFill("solid", fgColor="2E4057")
    c.font   = Font(color=C_GOLD, bold=True, name="Calibri", size=11)
    c.number_format = "+#,##0;-#,##0;0"
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border()

    ws.freeze_panes = "A3"

    # Mise en forme conditionnelle P&L
    ws.conditional_formatting.add("K3:K30", FormulaRule(
        formula=["K3>0"], fill=PatternFill("solid", fgColor="CCFFCC"),
        font=Font(bold=True, color="006400", name="Calibri", size=10)))
    ws.conditional_formatting.add("K3:K30", FormulaRule(
        formula=["K3<0"], fill=PatternFill("solid", fgColor="FFCCCC"),
        font=Font(bold=True, color=C_RED, name="Calibri", size=10)))

    return ws

# ─── Feuille 3 : Guide ────────────────────────────────────────────────────────

def make_guide(wb):
    ws = wb.create_sheet("📖 Guide", 2)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 55

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "Guide d'utilisation — BRVM Suivi de Portefeuille"
    t.fill  = PatternFill("solid", fgColor=C_HEADER)
    t.font  = Font(color="FFFFFF", bold=True, name="Calibri", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    sections = [
        ("WORKFLOW QUOTIDIEN", None, None),
        ("Étape 1", "10h15 GMT", "Reçois l'alerte Telegram (signal MPR/OBI ou top 3 diagnostic)"),
        ("Étape 2", "Si signal MEDIUM ou HIGH", "Va dans onglet 📊 Positions → remplis colonnes A à F"),
        ("Étape 3", "Pendant la session", "Mets à jour col K (Prix actuel) → le Statut se met à jour auto"),
        ("Étape 4", "Si ✅ Objectif atteint", "Vends le titre → note prix sortie dans Historique"),
        ("Étape 5", "Si 🛑 STOP LOSS", "Vends IMMÉDIATEMENT → perte limitée à 3% du capital investi"),
        ("", None, None),
        ("COLONNES À REMPLIR MANUELLEMENT", None, None),
        ("Col A", "Date entrée", "Date à laquelle tu as passé l'ordre d'achat"),
        ("Col B", "Ticker", "Symbole BRVM (ex : SGBC, ORAC, SNTS)"),
        ("Col C", "Nom société", "Nom (copier depuis l'alerte Telegram)"),
        ("Col D", "Signal", "H = HIGH (🔴), M = MEDIUM (🟠), L = LOW (🟡)"),
        ("Col E", "Qtité achetée", "Nombre de titres réellement achetés"),
        ("Col F", "Prix entrée", "Prix auquel tu as acheté (cours fixing confirmé)"),
        ("Col K", "Prix actuel", "⚡ À mettre à jour manuellement ou via Power Query"),
        ("", None, None),
        ("FORMULES AUTOMATIQUES", None, None),
        ("Col G", "Coût total", "= Qtité × Prix entrée"),
        ("Col H", "Stop-loss", "= Prix entrée × 0.97 (−3%)  → NE PAS DÉPASSER"),
        ("Col I", "Objectif", "= Prix entrée × 1.04 (+4%)"),
        ("Col J", "Max BRVM", "= Prix entrée × 1.075 (+7.5% = limite journalière BRVM)"),
        ("Col L", "P&L FCFA", "= (Prix actuel − Prix entrée) × Qtité"),
        ("Col M", "P&L %", "= (Prix actuel − Prix entrée) / Prix entrée"),
        ("Col N", "Statut", "Calculé automatiquement — guide ton action"),
        ("", None, None),
        ("RÈGLES DE GESTION", None, None),
        ("Budget", "75 000 FCFA", "80% max investi par signal (20% toujours en réserve)"),
        ("Stop-loss", "−3%", "Règle absolue : vendre si Statut = 🛑 STOP LOSS"),
        ("Objectif", "+4%", "Cible normale dans la session de fixing"),
        ("Max journalier", "+7.5%", "Limite BRVM — cours bloqué au-delà ce jour"),
        ("Signaux LOW", "Surveiller", "1 signal = observer seulement, ne pas acheter seul"),
        ("", None, None),
        ("ARCHIVAGE", None, None),
        ("Clôturer un trade", "→ Historique", "Copie la ligne de Positions vers 📁 Historique (valeurs, pas formules)"),
        ("Effacer la ligne", "après copie", "Efface la ligne dans Positions pour libérer la place"),
    ]

    row = 3
    for items in sections:
        ws.row_dimensions[row].height = 18
        if items[1] is None:
            ws.merge_cells(f"A{row}:C{row}")
            c = ws[f"A{row}"]
            c.value = items[0]
            if items[0]:
                c.fill  = PatternFill("solid", fgColor="2E75B6")
                c.font  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = border()
        else:
            for col, val in enumerate(items, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.font  = Font(name="Calibri", size=10,
                               bold=(col == 2), color="000000")
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=(col == 3))
                c.border = border()
                c.fill  = PatternFill("solid", fgColor=C_LIGHT if row % 2 == 0 else C_WHITE)
        row += 1

    return ws

# ─── Main ─────────────────────────────────────────────────────────────────────

wb = Workbook()
# Supprimer la feuille vide par défaut
del wb[wb.sheetnames[0]]

make_positions(wb)
make_historique(wb)
make_guide(wb)

out = "BRVM_Portefeuille.xlsx"
wb.save(out)
print(f"✅ Fichier créé : {out}")
print("   → Onglet '📊 Positions'  : saisir tes ordres + prix actuel")
print("   → Onglet '📁 Historique' : archiver les trades clôturés")
print("   → Onglet '📖 Guide'      : mode d'emploi complet")
