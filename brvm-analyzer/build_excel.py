"""
Génère etudes_actions_v2.xlsx — version améliorée de la fiche d'analyse BRVM.
Usage: python3 build_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import unicodedata, re

# ── Sanitize Excel formula strings (replace non-ASCII) ───────────────────────
_ACCENT_MAP = str.maketrans({
    'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
    'à': 'a', 'â': 'a', 'ä': 'a',
    'î': 'i', 'ï': 'i',
    'ô': 'o', 'ö': 'o',
    'ù': 'u', 'û': 'u', 'ü': 'u',
    'ç': 'c',
    'É': 'E', 'È': 'E', 'Ê': 'E',
    'À': 'A', 'Â': 'A',
    'Î': 'I', 'Ô': 'O', 'Ù': 'U', 'Û': 'U', 'Ç': 'C',
    '—': '-', '–': '-', '−': '-',
    '×': 'x',
    '≥': '>=', '≤': '<=',
    '➜': '->', '→': '->',
    '✅': '', '⚠': '', '🚫': '',
    '₀': '0', '₁': '1',
    '​': '',  # zero-width space
})

def sf(formula: str) -> str:
    """Sanitize a formula string: replace non-ASCII chars that break Excel."""
    if not formula or not formula.startswith('='):
        return formula
    return formula.translate(_ACCENT_MAP)

# ── helpers ──────────────────────────────────────────────────────────────────

def px(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def ft(bold=False, italic=False, color="000000", size=10):
    return Font(bold=bold, italic=italic, color=color, size=size, name="Calibri")

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bd(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bd_bottom(style="medium", color="2F5496"):
    s = Side(style=style, color=color)
    return Border(bottom=s)

NAVY   = "1F3864"
BLUE   = "2F5496"
LBLUE  = "D6E4F7"
ORANGE = "FF6600"
LORANGE= "FFF2CC"
GREEN  = "00B050"
LGREEN = "E2EFDA"
RED    = "FF0000"
LRED   = "FFDCE1"
GREY   = "F2F2F2"
WHITE  = "FFFFFF"
GOLD   = "C9AB2B"
DARK   = "404040"

def style_title(ws, row, col, text, merge_to_col=8,
                bg=NAVY, fg=WHITE, size=12):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = px(bg)
    cell.font = ft(bold=True, color=fg, size=size)
    cell.alignment = al("center", "center")
    if merge_to_col > col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=merge_to_col
        )
    ws.row_dimensions[row].height = 20

def style_section(ws, row, col, text, merge_to_col=8, bg=BLUE, fg=WHITE):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = px(bg)
    cell.font = ft(bold=True, color=fg, size=10)
    cell.alignment = al("left", "center")
    if merge_to_col > col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=merge_to_col
        )
    ws.row_dimensions[row].height = 18

def style_header(ws, row, cols, texts, bg=LBLUE):
    for col, text in zip(cols, texts):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = px(bg)
        cell.font = ft(bold=True, color=NAVY, size=9)
        cell.alignment = al("center", "center")
        cell.border = bd()

def style_label(ws, row, col, text, bg=GREY, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = px(bg)
    cell.font = ft(bold=bold, color=DARK, size=9)
    cell.alignment = al("left", "center")
    cell.border = bd()

def style_value(ws, row, col, value, fmt=None, bg=WHITE, bold=False, color=DARK):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = px(bg)
    cell.font = ft(bold=bold, color=color, size=9)
    cell.alignment = al("right", "center")
    cell.border = bd()
    if fmt:
        cell.number_format = fmt
    return cell

def style_formula(ws, row, col, formula, fmt=None, bg=WHITE, bold=False, color=DARK):
    cell = ws.cell(row=row, column=col, value=sf(formula))
    cell.fill = px(bg)
    cell.font = ft(bold=bold, color=color, size=9)
    cell.alignment = al("right", "center")
    cell.border = bd()
    if fmt:
        cell.number_format = fmt
    return cell

def blank_row(ws, row, height=6):
    ws.row_dimensions[row].height = height

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ══════════════════════════════════════════════════════════════════════════════
#  FEUILLE 1 : ETUDE
# ══════════════════════════════════════════════════════════════════════════════

def build_etude(ws):
    set_col_widths(ws, {
        1: 34, 2: 12, 3: 12, 4: 12,
        5: 12, 6: 10, 7: 14, 8: 12
    })

    r = 1

    # ── Titre principal ──────────────────────────────────────────────────────
    style_title(ws, r, 1,
        "FICHE D'ANALYSE BRVM — [NOM SOCIÉTÉ]   |   Date : [JJ/MM/AAAA]",
        merge_to_col=8, size=13)
    r += 1
    ws.cell(row=r, column=1,
        value="⚠ Remplacer [NOM SOCIÉTÉ] et [JJ/MM/AAAA] en cellule A1 avant utilisation"
    ).font = ft(italic=True, color="FF0000", size=8)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    blank_row(ws, r)
    r += 1

    # ════════════════════════════════════════════════════════════════════════
    #  BLOC A — ANALYSE FONDAMENTALE
    # ════════════════════════════════════════════════════════════════════════
    style_title(ws, r, 1, "A — ANALYSE FONDAMENTALE",
                merge_to_col=8, bg=NAVY, size=11)
    r += 1
    blank_row(ws, r); r += 1

    # ── A1 : Chiffre d'affaires / PNB ───────────────────────────────────────
    style_section(ws, r, 1, "A1 — CHIFFRE D'AFFAIRES ou PRODUIT NET BANCAIRE (FCFA)")
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5, 6, 7, 8],
                 ["Année", "N-4", "N-3", "N-2", "N-1", "N (dernier)",
                  "Croissance N/N-4 %", "TCAM 4 ans %"])
    CA_ROW = r + 1
    r += 1
    ws.cell(row=r, column=1, value="CA / PNB").fill = px(GREY)
    ws.cell(row=r, column=1).font = ft(size=9); ws.cell(row=r, column=1).border = bd()
    for c in [2, 3, 4, 5, 6]:
        style_value(ws, r, c, None, fmt='#,##0', bg=WHITE)
    # Croissance = (N - N-4) / N-4 × 100  ← formule corrigée
    style_formula(ws, r, 7, f"=IF(B{r}=0,\"\",((F{r}-B{r})/B{r})*100)",
                  fmt="0.00", bg=LGREEN)
    # TCAM = (N/N-4)^(1/4) - 1
    style_formula(ws, r, 8, f"=IF(B{r}=0,\"\",(F{r}/B{r})^(1/4)-1)",
                  fmt="0.00%", bg=LGREEN)
    CA_DATA_ROW = r
    r += 1
    blank_row(ws, r); r += 1

    # ── A2 : Résultat Net ────────────────────────────────────────────────────
    style_section(ws, r, 1, "A2 — RÉSULTAT NET (FCFA)")
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5, 6, 7, 8],
                 ["Année", "N-4", "N-3", "N-2", "N-1", "N (dernier)",
                  "Croissance N/N-4 %", "TCAM 4 ans %"])
    RN_ROW = r + 1
    r += 1
    ws.cell(row=r, column=1, value="Résultat Net").fill = px(GREY)
    ws.cell(row=r, column=1).font = ft(size=9); ws.cell(row=r, column=1).border = bd()
    for c in [2, 3, 4, 5, 6]:
        style_value(ws, r, c, None, fmt='#,##0', bg=WHITE)
    style_formula(ws, r, 7,
                  f"=IF(B{r}=0,\"\",((F{r}-B{r})/ABS(B{r}))*100)",
                  fmt="0.00", bg=LGREEN)
    style_formula(ws, r, 8,
                  f"=IF(OR(B{r}=0,F{r}/B{r}<0),\"N/A\",(F{r}/B{r})^(1/4)-1)",
                  fmt="0.00%", bg=LGREEN)
    RN_DATA_ROW = r
    r += 1
    blank_row(ws, r); r += 1

    # ── A3 : Marge Nette ────────────────────────────────────────────────────
    style_section(ws, r, 1, "A3 — MARGE NETTE (RN / CA × 100) — 5 dernières années")
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5, 6, 7],
                 ["Année", "RN (M FCFA)", "CA/PNB (M FCFA)",
                  "Marge %", "Interprétation",
                  "Seuil recommandé", "Moyenne 5 ans %"])
    MN_HEADER = r
    r += 1
    MARGE_ROWS = []
    annees = ["N-4", "N-3", "N-2", "N-1", "N"]
    for i, annee in enumerate(annees):
        ws.cell(row=r, column=1, value=annee).font = ft(size=9)
        ws.cell(row=r, column=1).fill = px(GREY); ws.cell(row=r, column=1).border = bd()
        style_value(ws, r, 2, None, fmt='#,##0', bg=WHITE)
        style_value(ws, r, 3, None, fmt='#,##0', bg=WHITE)
        style_formula(ws, r, 4,
                      f"=IFERROR(B{r}/C{r}*100,\"\")", fmt="0.00", bg=LGREEN)
        # Interprétation automatique
        style_formula(ws, r, 5,
                      f'=IFERROR(IF(D{r}>=10,"Très rentable",'
                      f'IF(D{r}>=5,"Rentable",'
                      f'IF(D{r}>=2,"Faiblement rentable","Très faible"))),"—")',
                      bg=WHITE)
        ws.cell(row=r, column=5).alignment = al("center", "center")
        ws.cell(row=r, column=5).font = ft(size=9)
        if i == 0:
            style_value(ws, r, 6, "> 5% = bon", bg=LORANGE)
            ws.cell(row=r, column=6).alignment = al("center", "center")
            style_formula(ws, r, 7,
                          f"=IFERROR(AVERAGE(D{r}:D{r+4}),\"\")",
                          fmt="0.00", bg=LGREEN, bold=True)
        else:
            for c in [6, 7]:
                ws.cell(row=r, column=c).border = bd()
        MARGE_ROWS.append(r)
        r += 1
    blank_row(ws, r); r += 1

    # ── A4 : Indicateurs de Valorisation ────────────────────────────────────
    style_section(ws, r, 1, "A4 — INDICATEURS DE VALORISATION")
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5],
                 ["Indicateur", "Formule", "Valeur calculée",
                  "Zone normale BRVM", "Signal"])
    VALO_HEADER = r
    r += 1

    # Saisies
    PRIX_ROW = r
    style_label(ws, r, 1, "Prix actuel de l'action (FCFA)")
    style_value(ws, r, 2, None, fmt='#,##0', bg=LORANGE)
    ws.cell(row=r, column=2).font = ft(bold=True, color=ORANGE, size=10)
    for c in [3, 4, 5]: ws.cell(row=r, column=c).border = bd()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws.cell(row=r, column=3, value="← Saisir ici le prix actuel").font = ft(italic=True, color="808080", size=8)
    r += 1

    BNPA_ROW = r
    style_label(ws, r, 1, "BNPA — Bénéfice Net Par Action (FCFA)")
    style_value(ws, r, 2, None, fmt='#,##0.00', bg=LORANGE)
    for c in [3, 4, 5]: ws.cell(row=r, column=c).border = bd()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 1

    CPTX_ROW = r
    style_label(ws, r, 1, "Capitaux Propres (FCFA)")
    style_value(ws, r, 2, None, fmt='#,##0', bg=LORANGE)
    for c in [3, 4, 5]: ws.cell(row=r, column=c).border = bd()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 1

    VALO_ENT_ROW = r
    style_label(ws, r, 1, "Valorisation / Capitalisation boursière (FCFA)")
    style_value(ws, r, 2, None, fmt='#,##0', bg=LORANGE)
    for c in [3, 4, 5]: ws.cell(row=r, column=c).border = bd()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 1

    NB_TITRES_ROW = r
    style_label(ws, r, 1, "Nombre de titres en circulation")
    style_value(ws, r, 2, None, fmt='#,##0', bg=LORANGE)
    for c in [3, 4, 5]: ws.cell(row=r, column=c).border = bd()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 1

    blank_row(ws, r, 4); r += 1

    # Ratios calculés
    ratios = [
        ("PER — Price Earning Ratio",
         f"=B{PRIX_ROW}/B{BNPA_ROW}",
         "9 – 15",
         f'=IFERROR(IF(C{r}>15,"Surévalué",IF(C{r}<9,"Sous-évalué","Zone normale")),"—")'),
        ("VMC — Valeur Mathématique Comptable (FCFA)",
         f"=B{CPTX_ROW}/B{NB_TITRES_ROW}",
         "—",
         f'=IFERROR(IF(B{PRIX_ROW}>C{r},"Action surévaluée","Action sous-évaluée"),"—")'),
        ("PBR — Price to Book Ratio",
         f"=B{VALO_ENT_ROW}/B{CPTX_ROW}",
         "< 1 = bon",
         f'=IFERROR(IF(C{r}<1,"Décote comptable",IF(C{r}<2,"Normal","Prime élevée")),"—")'),
        ("ROE — Return on Equity (%)",
         f"=IFERROR(B{RN_DATA_ROW}*1000000/B{CPTX_ROW}*100,\"\")",
         "> 10% = bon",
         f'=IFERROR(IF(C{r}>=15,"Excellent",IF(C{r}>=10,"Bon",IF(C{r}>=5,"Moyen","Faible"))),"—")'),
        ("RVC = PER × PBR",
         f"=IFERROR(C{r-4}*C{r-2},\"\")",
         "< 22 = Décote",
         f'=IFERROR(IF(C{r}<22,"Décote (achat possible)","Surévaluation"),"—")'),
    ]

    PER_ROW = r
    VMC_ROW = r + 1
    PBR_ROW = r + 2
    ROE_ROW = r + 3
    RVC_ROW = r + 4

    for label, formula_val, zone, signal_formula in ratios:
        style_label(ws, r, 1, label)
        style_value(ws, r, 2, formula_val, fmt='0.00', bg=LGREEN)
        ws.cell(row=r, column=2).value = formula_val
        ws.cell(row=r, column=2).number_format = "0.00"
        # Valeur calculée : formule dynamique
        style_formula(ws, r, 3, formula_val, fmt='0.00', bg=LGREEN)
        style_value(ws, r, 4, zone, bg=LORANGE)
        ws.cell(row=r, column=4).alignment = al("center", "center")
        style_formula(ws, r, 5, signal_formula, bg=WHITE)
        ws.cell(row=r, column=5).alignment = al("center", "center")
        ws.cell(row=r, column=5).font = ft(size=9)
        r += 1
    blank_row(ws, r); r += 1

    # ── A5 : Taux de distribution & Dividendes ───────────────────────────────
    style_section(ws, r, 1, "A5 — DIVIDENDES & TAUX DE DISTRIBUTION")
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5, 6],
                 ["Année", "Dividende brut (FCFA)", "BNPA (FCFA)",
                  "Taux distribution %", "Rendement %", "Progression %"])
    DVD_HEADER = r
    DVD_ROWS = []
    r += 1
    for i in range(5):
        annee_label = ["N-4", "N-3", "N-2", "N-1", "N"][i]
        style_label(ws, r, 1, annee_label)
        style_value(ws, r, 2, None, fmt='#,##0.00', bg=WHITE)
        style_value(ws, r, 3, None, fmt='#,##0.00', bg=WHITE)
        # Taux distribution = DVD / BNPA × 100 (après précompte 15%)
        style_formula(ws, r, 4,
                      f"=IFERROR(B{r}/(C{r}*0.85)*100,\"\")",
                      fmt="0.00", bg=LGREEN)
        # Rendement = Dividende / Prix × 100
        style_formula(ws, r, 5,
                      f"=IFERROR(B{r}/B{PRIX_ROW}*100,\"\")",
                      fmt="0.00", bg=LGREEN)
        if i == 0:
            ws.cell(row=r, column=6).border = bd()
            ws.cell(row=r, column=6, value="—").font = ft(size=9)
            ws.cell(row=r, column=6).alignment = al("center", "center")
        else:
            style_formula(ws, r, 6,
                          f"=IFERROR((B{r}-B{r-1})/ABS(B{r-1})*100,\"\")",
                          fmt="0.00", bg=LGREEN)
        DVD_ROWS.append(r)
        r += 1

    # Total dividendes & rendement moyen
    style_label(ws, r, 1, "TOTAL dividendes & rendement moyen", bold=True)
    style_formula(ws, r, 2, f"=SUM(B{DVD_ROWS[0]}:B{DVD_ROWS[-1]})",
                  fmt='#,##0.00', bg=LGREEN, bold=True)
    for c in [3]: ws.cell(row=r, column=c).border = bd()
    style_formula(ws, r, 4,
                  f"=IFERROR(AVERAGE(D{DVD_ROWS[0]}:D{DVD_ROWS[-1]}),\"\")",
                  fmt="0.00", bg=LGREEN, bold=True)
    style_formula(ws, r, 5,
                  f"=IFERROR(AVERAGE(E{DVD_ROWS[0]}:E{DVD_ROWS[-1]}),\"\")",
                  fmt="0.00", bg=LGREEN, bold=True)
    for c in [6]: ws.cell(row=r, column=c).border = bd()
    DVD_TOTAL_ROW = r
    r += 1
    blank_row(ws, r); r += 1

    # ── A6 : Fonds Propres ──────────────────────────────────────────────────
    style_section(ws, r, 1, "A6 — FONDS PROPRES (FCFA)")
    r += 1
    style_header(ws, r, [1, 2, 3, 4],
                 ["Composante", "Année N-1", "Année N", "Variation %"])
    FP_HEADER = r
    r += 1
    fp_items = ["Capital souscrit", "Réserves", "Report à nouveau", "Résultat Net", "Dividendes versés"]
    FP_ROWS = []
    for item in fp_items:
        style_label(ws, r, 1, item)
        style_value(ws, r, 2, None, fmt='#,##0', bg=WHITE)
        style_value(ws, r, 3, None, fmt='#,##0', bg=WHITE)
        style_formula(ws, r, 4,
                      f"=IFERROR((C{r}-B{r})/ABS(B{r})*100,\"\")",
                      fmt="0.00", bg=LGREEN)
        FP_ROWS.append(r)
        r += 1

    # Total fonds propres
    style_label(ws, r, 1, "TOTAL FONDS PROPRES", bold=True)
    style_formula(ws, r, 2, f"=SUM(B{FP_ROWS[0]}:B{FP_ROWS[-1]})",
                  fmt='#,##0', bg=LGREEN, bold=True)
    style_formula(ws, r, 3, f"=SUM(C{FP_ROWS[0]}:C{FP_ROWS[-1]})",
                  fmt='#,##0', bg=LGREEN, bold=True)
    style_formula(ws, r, 4,
                  f"=IFERROR((C{r}-B{r})/ABS(B{r})*100,\"\")",
                  fmt="0.00", bg=LGREEN, bold=True)
    FP_TOTAL_ROW = r
    r += 1
    blank_row(ws, r); r += 1

    # ── A7 : Notation & Activités récentes ──────────────────────────────────
    style_section(ws, r, 1, "A7 — NOTATION FINANCIÈRE & ACTIVITÉS RÉCENTES")
    r += 1
    style_header(ws, r, [1, 2, 3, 4],
                 ["Indicateur", "Valeur", "Source", "Date rapport"])
    r += 1
    items_notation = [
        ("Notation court terme", "", "Agence de notation", ""),
        ("Notation long terme", "", "Agence de notation", ""),
        ("Variation CA/PNB (trimestre)", "", "RichBourse", ""),
        ("Variation Résultat Net (trimestre)", "", "RichBourse", ""),
    ]
    for label, val, src, date in items_notation:
        style_label(ws, r, 1, label)
        style_value(ws, r, 2, val or None, bg=WHITE)
        ws.cell(row=r, column=2).alignment = al("center", "center")
        style_value(ws, r, 3, src, bg=GREY)
        ws.cell(row=r, column=3).alignment = al("center", "center")
        style_value(ws, r, 4, date or None, bg=GREY)
        ws.cell(row=r, column=4).alignment = al("center", "center")
        r += 1
    blank_row(ws, r); r += 1

    # ── A8 : Comparaison Marché vs Épargne ──────────────────────────────────
    style_section(ws, r, 1, "A8 — COMPARAISON MARCHÉ BOURSIER VS ÉPARGNE")
    r += 1

    style_header(ws, r, [1, 2, 3],
                 ["Paramètre", "Marché Boursier", "Épargne (référence)"])
    r += 1

    COMP_ROWS = {}

    items_comp = [
        ("Montant investi (FCFA)", "#,##0", None),
        ("Prix d'achat (FCFA)", "#,##0", None),
        ("CMP (prix + 1,4% frais achat)", "#,##0.00", None),
        ("Nombre de titres acquis", "#,##0.00", None),
        ("Prix actuel titre (FCFA)", "#,##0", None),
        ("Prix net après taxe cession 1,4% (FCFA)", "#,##0.00", None),
        ("Plus-value unitaire (FCFA)", "#,##0.00", None),
        ("Plus-value totale (FCFA)", "#,##0", None),
        ("Total dividendes perçus (FCFA)", "#,##0", None),
        ("Rendement total (FCFA)", "#,##0", None),
        ("Performance (%)", "0.00", None),
    ]

    for label, fmt, _ in items_comp:
        style_label(ws, r, 1, label)
        style_value(ws, r, 2, None, fmt=fmt, bg=WHITE)
        style_value(ws, r, 3, None, fmt=fmt, bg=WHITE)
        COMP_ROWS[label] = r
        r += 1

    # Pré-remplir les formules de la colonne marché
    inv_r    = COMP_ROWS["Montant investi (FCFA)"]
    achat_r  = COMP_ROWS["Prix d'achat (FCFA)"]
    cmp_r    = COMP_ROWS["CMP (prix + 1,4% frais achat)"]
    titres_r = COMP_ROWS["Nombre de titres acquis"]
    prix_r   = COMP_ROWS["Prix actuel titre (FCFA)"]
    net_r    = COMP_ROWS["Prix net après taxe cession 1,4% (FCFA)"]
    pvu_r    = COMP_ROWS["Plus-value unitaire (FCFA)"]
    pvt_r    = COMP_ROWS["Plus-value totale (FCFA)"]
    dvd_r    = COMP_ROWS["Total dividendes perçus (FCFA)"]
    rdt_r    = COMP_ROWS["Rendement total (FCFA)"]
    perf_r   = COMP_ROWS["Performance (%)"]

    ws.cell(row=cmp_r,    column=2, value=f"=IFERROR(B{achat_r}*1.014,\"\")")
    ws.cell(row=titres_r, column=2, value=f"=IFERROR(B{inv_r}/B{cmp_r},\"\")")
    ws.cell(row=net_r,    column=2, value=f"=IFERROR(B{prix_r}-(B{prix_r}*0.014),\"\")")
    ws.cell(row=pvu_r,    column=2, value=f"=IFERROR(B{net_r}-B{cmp_r},\"\")")
    ws.cell(row=pvt_r,    column=2, value=f"=IFERROR(B{pvu_r}*B{titres_r},\"\")")
    ws.cell(row=dvd_r,    column=2, value=f"=IFERROR(B{titres_r}*C{dvd_r},\"\")")  # C = total dvd par titre
    ws.cell(row=rdt_r,    column=2, value=f"=IFERROR(B{pvt_r}+B{dvd_r},\"\")")
    ws.cell(row=perf_r,   column=2, value=f"=IFERROR(B{rdt_r}/B{inv_r}*100,\"\")")

    for row_n in [cmp_r, titres_r, net_r, pvu_r, pvt_r, dvd_r, rdt_r, perf_r]:
        ws.cell(row=row_n, column=2).fill = px(LGREEN)
        ws.cell(row=row_n, column=2).font = ft(size=9, color=DARK)
        ws.cell(row=row_n, column=2).alignment = al("right", "center")
        ws.cell(row=row_n, column=2).border = bd()

    ws.cell(row=dvd_r, column=3,
            value="← Total dividendes reçus par titre sur la période").font = ft(italic=True, color="808080", size=8)
    ws.cell(row=dvd_r, column=3).border = bd()

    # Verdict
    verdict_r = r
    style_label(ws, r, 1, "VERDICT", bold=True)
    style_formula(ws, r, 2,
                  f'=IFERROR(IF(B{perf_r}>C{perf_r},"MARCHE BOURSIER","EPARGNE"),"Saisir les données")',
                  bg=LGREEN, bold=True)
    ws.cell(row=r, column=2).alignment = al("center", "center")
    style_value(ws, r, 3, None, bg=WHITE)
    ws.cell(row=r, column=3, value="← Saisir le rendement épargne %").font = ft(italic=True, color="808080", size=8)
    r += 1
    blank_row(ws, r); r += 1

    # ── A9 : Conclusion Fondamentale ────────────────────────────────────────
    style_section(ws, r, 1, "A9 — CONCLUSION ANALYSE FONDAMENTALE (texte libre)")
    r += 1
    cell = ws.cell(row=r, column=1)
    cell.value = ("← Saisir ici ta synthèse fondamentale : "
                  "PER, valorisation, tendance CA/RN, politique dividendes, "
                  "fonds propres, comparaison épargne…")
    cell.font = ft(italic=True, color="808080", size=9)
    cell.fill = px(WHITE)
    cell.alignment = al("left", "center", wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=8)
    ws.row_dimensions[r].height = 50
    r += 3
    blank_row(ws, r); r += 1

    # ════════════════════════════════════════════════════════════════════════
    #  BLOC B — ANALYSE TECHNIQUE
    # ════════════════════════════════════════════════════════════════════════
    style_title(ws, r, 1, "B — ANALYSE TECHNIQUE",
                merge_to_col=8, bg=NAVY, size=11)
    r += 1
    blank_row(ws, r); r += 1

    # ── B1 : Indicateurs techniques ─────────────────────────────────────────
    style_section(ws, r, 1, "B1 — INDICATEURS TECHNIQUES (saisie manuelle depuis graphique)")
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5],
                 ["Indicateur", "Signal observé", "Valeur / Niveau",
                  "Points attribués", "Référence"])
    r += 1

    TECH_ROWS = []
    tech_items = [
        ("Moyenne Mobile 20 jours (MM20)",
         ["Zone de survente", "Neutre", "Zone de surachat"],
         "MM20 : prix en dessous = survente"),
        ("Bandes de Bollinger",
         ["Zone de survente", "Neutre", "Zone de surachat"],
         "Prix < bande inf = survente"),
        ("MACD",
         ["Zone de survente", "Neutre", "Zone de surachat"],
         "MACD > Signal = haussier"),
        ("RSI",
         ["Zone de survente (RSI < 30)", "Neutre (30-70)", "Zone de surachat (RSI > 70)"],
         "RSI < 30 = survente"),
        ("Volume (tendance récente)",
         ["Volume croissant haussier", "Neutre", "Volume décroissant"],
         "Volume confirme la tendance"),
    ]

    POINTS_CELLS = []
    for indicator, options, ref in tech_items:
        style_label(ws, r, 1, indicator)
        # Colonne B : liste déroulante (validation) — signal
        cell_sig = ws.cell(row=r, column=2)
        cell_sig.value = "← Saisir signal"
        cell_sig.font = ft(italic=True, color="808080", size=8)
        cell_sig.fill = px(WHITE)
        cell_sig.alignment = al("center", "center")
        cell_sig.border = bd()
        # Colonne C : valeur numérique
        style_value(ws, r, 3, None, fmt='0.00', bg=WHITE)
        # Colonne D : points (1 si survente, 0 si neutre, -1 si surachat)
        style_formula(ws, r, 4,
                      f'=IFERROR(IF(ISNUMBER(SEARCH("survente",B{r})),1,'
                      f'IF(ISNUMBER(SEARCH("surachat",B{r})),-1,0)),0)',
                      fmt='0', bg=LGREEN)
        POINTS_CELLS.append(r)
        # Colonne E : référence
        cell_ref = ws.cell(row=r, column=5, value=ref)
        cell_ref.font = ft(italic=True, color="606060", size=8)
        cell_ref.fill = px(GREY)
        cell_ref.alignment = al("left", "center", wrap=True)
        cell_ref.border = bd()
        TECH_ROWS.append(r)
        r += 1

    # Total points techniques
    SCORE_TECH_ROW = r
    style_label(ws, r, 1, "SCORE TECHNIQUE (max 5, min -5)", bold=True)
    for c in [2, 3]: ws.cell(row=r, column=c).border = bd()
    style_formula(ws, r, 4,
                  f"=SUM(D{POINTS_CELLS[0]}:D{POINTS_CELLS[-1]})",
                  fmt='0', bg=LGREEN, bold=True)
    style_formula(ws, r, 5,
                  f'=IFERROR(IF(D{r}>0,"Zone SURVENTE - Achat possible",'
                  f'IF(D{r}<0,"Zone SURACHAT - Vente possible","NEUTRE")),"—")',
                  bg=WHITE, bold=True)
    ws.cell(row=r, column=5).alignment = al("center", "center")
    ws.cell(row=r, column=5).font = ft(bold=True, size=9)
    r += 1
    blank_row(ws, r); r += 1

    # ── B2 : Prix Médians ───────────────────────────────────────────────────
    style_section(ws, r, 1, "B2 — PRIX MÉDIANS (PM)")
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5],
                 ["Horizon", "Plus Haut", "Plus Bas",
                  "Prix Médian", "Signal (vs prix actuel)"])
    r += 1

    PM_ROWS = []
    for horizon in ["52 semaines (1 an)", "2 ans"]:
        style_label(ws, r, 1, horizon)
        style_value(ws, r, 2, None, fmt='#,##0', bg=WHITE)   # Plus haut
        style_value(ws, r, 3, None, fmt='#,##0', bg=WHITE)   # Plus bas
        style_formula(ws, r, 4,
                      f"=IFERROR((B{r}+C{r})/2,\"\")",
                      fmt='#,##0.00', bg=LGREEN)
        style_formula(ws, r, 5,
                      f'=IFERROR(IF(B{PRIX_ROW}<D{r},"EN DESSOUS du PM - Survente","AU-DESSUS du PM - Surachat"),"—")',
                      bg=WHITE)
        ws.cell(row=r, column=5).alignment = al("center", "center")
        ws.cell(row=r, column=5).font = ft(size=9)
        PM_ROWS.append(r)
        r += 1
    blank_row(ws, r); r += 1

    # ── B3 : PER actualisé ──────────────────────────────────────────────────
    style_section(ws, r, 1, "B3 — PER ACTUALISÉ")
    r += 1
    style_header(ws, r, [1, 2, 3, 4],
                 ["Paramètre", "Valeur", "Source", "Note"])
    r += 1

    PER_ACT_ROWS = {}
    per_items = [
        ("Prix actuel (FCFA)",
         f"=B{PRIX_ROW}", "Section A4", "#,##0"),
        ("BNPA N (FCFA)",
         f"=B{BNPA_ROW}", "Section A4", "#,##0.00"),
        ("Évolution RN dernier trimestre (%)",
         None, "RichBourse / Rapport", "0.00"),
        ("BNPA actualisé (FCFA)",
         None, "= BNPA × (1 + évol%)", "#,##0.00"),
        ("PER actualisé",
         None, "= Prix / BNPA actualisé", "0.00"),
    ]

    for label, val, src, fmt in per_items:
        style_label(ws, r, 1, label)
        if val:
            style_formula(ws, r, 2, val, fmt=fmt, bg=LGREEN)
        else:
            style_value(ws, r, 2, None, fmt=fmt, bg=WHITE)
        style_value(ws, r, 3, src, bg=GREY)
        ws.cell(row=r, column=3).alignment = al("center", "center")
        ws.cell(row=r, column=4).border = bd()
        PER_ACT_ROWS[label] = r
        r += 1

    # Formules BNPA actualisé et PER actualisé
    bnpa_act_r = PER_ACT_ROWS["BNPA actualisé (FCFA)"]
    evol_r     = PER_ACT_ROWS["Évolution RN dernier trimestre (%)"]
    prix_act_r = PER_ACT_ROWS["Prix actuel (FCFA)"]
    bnpa_r_ref = PER_ACT_ROWS["BNPA N (FCFA)"]
    per_act_r  = PER_ACT_ROWS["PER actualisé"]

    ws.cell(row=bnpa_act_r, column=2,
            value=f"=IFERROR(B{bnpa_r_ref}*(1+B{evol_r}/100),\"\")")
    ws.cell(row=bnpa_act_r, column=2).fill = px(LGREEN)
    ws.cell(row=bnpa_act_r, column=2).font = ft(size=9)
    ws.cell(row=bnpa_act_r, column=2).alignment = al("right", "center")
    ws.cell(row=bnpa_act_r, column=2).border = bd()
    ws.cell(row=bnpa_act_r, column=2).number_format = "#,##0.00"

    ws.cell(row=per_act_r, column=2,
            value=f"=IFERROR(B{prix_act_r}/B{bnpa_act_r},\"\")")
    ws.cell(row=per_act_r, column=2).fill = px(LGREEN)
    ws.cell(row=per_act_r, column=2).font = ft(size=9)
    ws.cell(row=per_act_r, column=2).alignment = al("right", "center")
    ws.cell(row=per_act_r, column=2).border = bd()
    ws.cell(row=per_act_r, column=2).number_format = "0.00"

    # Signal PER
    style_formula(ws, r-1, 4,
                  f'=IFERROR(IF(B{per_act_r}>15,"Surévalué — Surachat",'
                  f'IF(B{per_act_r}<9,"Sous-évalué — Survente","Normal (9–15)")),"—")',
                  bg=WHITE)
    ws.cell(row=per_act_r, column=4).alignment = al("center", "center")
    ws.cell(row=per_act_r, column=4).font = ft(size=9)
    blank_row(ws, r); r += 1

    # ── B4 : Prix Cible aux Dividendes (PCD) ────────────────────────────────
    style_section(ws, r, 1, "B4 — PRIX CIBLE AUX DIVIDENDES (PCD)")
    r += 1

    note_pcd = ws.cell(row=r, column=1,
        value=("Note : Le PCD n'est pertinent que si l'entreprise distribue des dividendes. "
               "Si dividende = 0, utiliser la section Gordon-Shapiro (B5)."))
    note_pcd.font = ft(italic=True, color=ORANGE, size=8)
    note_pcd.fill = px(LORANGE)
    note_pcd.alignment = al("left", "center", wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 25
    r += 1

    style_header(ws, r, [1, 2, 3, 4],
                 ["Paramètre", "Valeur", "Formule", "Note"])
    r += 1

    PCD_ROWS = {}
    pcd_items = [
        ("Évolution RN (%)", None, "← même valeur qu'en B3", "0.00"),
        ("Dividende brut N (FCFA)", None, "← saisir le dernier dividende", "#,##0.00"),
        ("Dividende prévisionnel (FCFA)", None, "= DVD × (1 + évol%)", "#,##0.00"),
        ("Rendement sectoriel BRVM (%)", 6.5, "Référence : 5–8%", "0.00"),
        ("PCD = 100 × DVD_prev / Rendement", None, "= 100 × Dprev / Rend%", "#,##0.00"),
        ("Zone cible basse (−5%)", None, "= PCD × 0.95", "#,##0.00"),
        ("Zone cible haute (+5%)", None, "= PCD × 1.05", "#,##0.00"),
        ("Signal (prix actuel vs zone cible)", None, "Auto", ""),
    ]

    for label, val, note, fmt in pcd_items:
        style_label(ws, r, 1, label)
        if val is not None:
            style_value(ws, r, 2, val, fmt=fmt, bg=LORANGE)
        else:
            style_value(ws, r, 2, None, fmt=fmt, bg=WHITE)
        style_value(ws, r, 3, note, bg=GREY)
        ws.cell(row=r, column=3).alignment = al("center", "center")
        ws.cell(row=r, column=4).border = bd()
        PCD_ROWS[label] = r
        r += 1

    evol_pcd_r = PCD_ROWS["Évolution RN (%)"]
    dvd_n_r    = PCD_ROWS["Dividende brut N (FCFA)"]
    dvd_prev_r = PCD_ROWS["Dividende prévisionnel (FCFA)"]
    rend_r     = PCD_ROWS["Rendement sectoriel BRVM (%)"]
    pcd_r      = PCD_ROWS["PCD = 100 × DVD_prev / Rendement"]
    zone_bas_r = PCD_ROWS["Zone cible basse (−5%)"]
    zone_haut_r= PCD_ROWS["Zone cible haute (+5%)"]
    signal_pcd_r = PCD_ROWS["Signal (prix actuel vs zone cible)"]

    for frow, formula, fmt in [
        (dvd_prev_r,
         f"=IFERROR(IF(B{dvd_n_r}=0,\"Aucun dividende\",B{dvd_n_r}*(1+B{evol_pcd_r}/100)),\"\")",
         "#,##0.00"),
        (pcd_r,
         f"=IFERROR(IF(B{dvd_n_r}=0,\"N/A\",100*B{dvd_prev_r}/B{rend_r}),\"\")",
         "#,##0.00"),
        (zone_bas_r,
         f"=IFERROR(IF(ISNUMBER(B{pcd_r}),B{pcd_r}*0.95,\"N/A\"),\"\")",
         "#,##0.00"),
        (zone_haut_r,
         f"=IFERROR(IF(ISNUMBER(B{pcd_r}),B{pcd_r}*1.05,\"N/A\"),\"\")",
         "#,##0.00"),
        (signal_pcd_r,
         f'=IFERROR(IF(NOT(ISNUMBER(B{pcd_r})),"PCD N/A (pas de dividende)",'
         f'IF(B{PRIX_ROW}<B{zone_bas_r},"EN DESSOUS zone cible - Survente",'
         f'IF(B{PRIX_ROW}>B{zone_haut_r},"AU-DESSUS zone cible - Surachat","DANS LA ZONE"))),\"\")',
         ""),
    ]:
        ws.cell(row=frow, column=2, value=sf(formula))
        ws.cell(row=frow, column=2).fill = px(LGREEN)
        ws.cell(row=frow, column=2).font = ft(size=9)
        ws.cell(row=frow, column=2).alignment = al("right", "center")
        ws.cell(row=frow, column=2).border = bd()
        if fmt:
            ws.cell(row=frow, column=2).number_format = fmt

    blank_row(ws, r); r += 1

    # ── B5 : Gordon-Shapiro ─────────────────────────────────────────────────
    style_section(ws, r, 1,
                  "B5 — MODÈLE DE GORDON-SHAPIRO (Prix d'entrée théorique)")
    r += 1
    note_gs = ws.cell(row=r, column=1,
        value="P₀ = D₁ / (r − g)   |   D₁ = dividende prévu, r = taux exigé, g = taux croissance dividendes")
    note_gs.font = ft(italic=True, color=NAVY, size=9)
    note_gs.fill = px(LBLUE)
    note_gs.alignment = al("center", "center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    style_header(ws, r, [1, 2, 3, 4],
                 ["Paramètre", "Valeur", "Formule", "Note"])
    r += 1

    GS_ROWS = {}
    gs_items = [
        ("D₁ — Dividende attendu l'an prochain (FCFA)", None,
         "← Hypothèse ou dernier DVD × (1+g)", "#,##0.00"),
        ("r — Taux de rentabilité exigé (%)", 8.0,
         "Coût du capital / rendement cible", "0.00"),
        ("g — Taux de croissance dividendes (%)", 2.0,
         "Croissance annuelle supposée", "0.00"),
        ("P₀ = D₁ / (r − g)  [Prix théorique FCFA]", None,
         "Auto-calculé", "#,##0.00"),
        ("Signal (prix actuel vs P₀)", None,
         "Auto", ""),
    ]

    for label, val, note, fmt in gs_items:
        style_label(ws, r, 1, label)
        if val is not None:
            style_value(ws, r, 2, val, fmt=fmt, bg=LORANGE)
        else:
            style_value(ws, r, 2, None, fmt=fmt, bg=WHITE)
        style_value(ws, r, 3, note, bg=GREY)
        ws.cell(row=r, column=3).alignment = al("center", "center")
        ws.cell(row=r, column=4).border = bd()
        GS_ROWS[label] = r
        r += 1

    d1_r  = GS_ROWS["D₁ — Dividende attendu l'an prochain (FCFA)"]
    rr_r  = GS_ROWS["r — Taux de rentabilité exigé (%)"]
    g_r   = GS_ROWS["g — Taux de croissance dividendes (%)"]
    p0_r  = GS_ROWS["P₀ = D₁ / (r − g)  [Prix théorique FCFA]"]
    sig_r = GS_ROWS["Signal (prix actuel vs P₀)"]

    ws.cell(row=p0_r, column=2,
            value=f"=IFERROR(IF(B{d1_r}=0,\"D1 non renseigne\","
                  f"B{d1_r}/((B{rr_r}/100)-(B{g_r}/100))),\"\")")
    ws.cell(row=p0_r, column=2).fill = px(LGREEN)
    ws.cell(row=p0_r, column=2).font = ft(bold=True, size=9)
    ws.cell(row=p0_r, column=2).alignment = al("right", "center")
    ws.cell(row=p0_r, column=2).border = bd()
    ws.cell(row=p0_r, column=2).number_format = "#,##0.00"

    ws.cell(row=sig_r, column=2,
            value=f'=IFERROR(IF(NOT(ISNUMBER(B{p0_r})),"Renseigner D1",'
                  f'IF(B{PRIX_ROW}<B{p0_r}*0.95,"Prix sous P0 - Opportunite achat",'
                  f'IF(B{PRIX_ROW}>B{p0_r}*1.05,"Prix sur P0 - Surevalue","Dans la zone"))),"")')
    ws.cell(row=sig_r, column=2).fill = px(LGREEN)
    ws.cell(row=sig_r, column=2).font = ft(size=9)
    ws.cell(row=sig_r, column=2).alignment = al("center", "center")
    ws.cell(row=sig_r, column=2).border = bd()

    blank_row(ws, r); r += 1

    # ── B6 : Stop-Loss & Risk/Reward ────────────────────────────────────────
    style_section(ws, r, 1, "B6 — GESTION DU RISQUE : STOP-LOSS & RISK/REWARD")
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5],
                 ["Paramètre", "Valeur", "Formule", "Note", "Signal"])
    r += 1

    SL_ROWS = {}
    sl_items = [
        ("Prix d'achat visé (FCFA)", None, "← Saisir", "#,##0"),
        ("Stop-Loss (FCFA) — perte max 10%", None, "= Prix × 0.90", "#,##0.00"),
        ("Objectif de prix (Take Profit) (FCFA)", None, "← Saisir objectif", "#,##0"),
        ("Perte potentielle (FCFA/titre)", None, "= Prix − Stop", "#,##0.00"),
        ("Gain potentiel (FCFA/titre)", None, "= Objectif − Prix", "#,##0.00"),
        ("Ratio Risk/Reward", None, "= Gain / Perte", "0.00"),
    ]

    for label, val, note, fmt in sl_items:
        style_label(ws, r, 1, label)
        if val is not None:
            style_value(ws, r, 2, val, fmt=fmt, bg=LORANGE)
        else:
            style_value(ws, r, 2, None, fmt=fmt, bg=WHITE)
        style_value(ws, r, 3, note, bg=GREY)
        ws.cell(row=r, column=3).alignment = al("center", "center")
        ws.cell(row=r, column=4).border = bd()
        ws.cell(row=r, column=5).border = bd()
        SL_ROWS[label] = r
        r += 1

    pa_r  = SL_ROWS["Prix d'achat visé (FCFA)"]
    sl_r  = SL_ROWS["Stop-Loss (FCFA) — perte max 10%"]
    tp_r  = SL_ROWS["Objectif de prix (Take Profit) (FCFA)"]
    pp_r  = SL_ROWS["Perte potentielle (FCFA/titre)"]
    gp_r  = SL_ROWS["Gain potentiel (FCFA/titre)"]
    rr_sl_r = SL_ROWS["Ratio Risk/Reward"]

    for frow, formula, fmt, signal in [
        (sl_r,    f"=IFERROR(B{pa_r}*0.90,\"\")", "#,##0.00", ""),
        (pp_r,    f"=IFERROR(B{pa_r}-B{sl_r},\"\")", "#,##0.00", ""),
        (gp_r,    f"=IFERROR(B{tp_r}-B{pa_r},\"\")", "#,##0.00", ""),
        (rr_sl_r, f"=IFERROR(B{gp_r}/B{pp_r},\"\")", "0.00",
         f'=IFERROR(IF(B{rr_sl_r}>=2,"Bon (>= 2:1)",'
         f'IF(B{rr_sl_r}>=1,"Acceptable (1:1)","Mauvais")),"")'),
    ]:
        ws.cell(row=frow, column=2, value=formula)
        ws.cell(row=frow, column=2).fill = px(LGREEN)
        ws.cell(row=frow, column=2).font = ft(size=9)
        ws.cell(row=frow, column=2).alignment = al("right", "center")
        ws.cell(row=frow, column=2).border = bd()
        ws.cell(row=frow, column=2).number_format = fmt
        if signal:
            ws.cell(row=frow, column=5, value=signal)
            ws.cell(row=frow, column=5).fill = px(LGREEN)
            ws.cell(row=frow, column=5).font = ft(size=9)
            ws.cell(row=frow, column=5).alignment = al("center", "center")
            ws.cell(row=frow, column=5).border = bd()

    blank_row(ws, r); r += 1

    # ── B7 : Conclusion Technique ────────────────────────────────────────────
    style_section(ws, r, 1, "B7 — CONCLUSION ANALYSE TECHNIQUE (texte libre)")
    r += 1
    cell = ws.cell(row=r, column=1)
    cell.value = "← Saisir ici ta synthèse technique : indicateurs, zone de prix, recommandation d'achat par palier…"
    cell.font = ft(italic=True, color="808080", size=9)
    cell.fill = px(WHITE)
    cell.alignment = al("left", "center", wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=8)
    ws.row_dimensions[r].height = 50
    r += 3


# ══════════════════════════════════════════════════════════════════════════════
#  FEUILLE 2 : PROFIL
# ══════════════════════════════════════════════════════════════════════════════

def build_profil(ws):
    set_col_widths(ws, {1: 30, 2: 25, 3: 30, 4: 25})

    r = 1
    style_title(ws, r, 1, "PROFIL DE LA SOCIÉTÉ", merge_to_col=4, size=13)
    r += 1
    blank_row(ws, r); r += 1

    sections = [
        ("IDENTIFICATION", [
            ("Nom de la société", ""),
            ("Ticker BRVM", ""),
            ("Secteur d'activité", ""),
            ("Pays d'origine", ""),
            ("Date d'introduction en bourse", ""),
            ("Marché", "BRVM — Bourse Régionale des Valeurs Mobilières"),
        ]),
        ("CAPITAL & ACTIONNARIAT", [
            ("Capital social (FCFA)", ""),
            ("Nombre de titres", ""),
            ("Prix d'émission initial (FCFA)", ""),
            ("Flottant (%)", ""),
            ("Principaux actionnaires", ""),
        ]),
        ("LIQUIDITÉ DU TITRE", [
            ("Volume moyen journalier (titres)", "← Source : BRVM.org / RichBourse"),
            ("Volume moyen journalier (FCFA)", ""),
            ("Fréquence de cotation (%)", "← % de séances avec transaction"),
            ("Appréciation de la liquidité", "← Forte / Moyenne / Faible"),
        ]),
        ("NOTATION & CONFORMITÉ", [
            ("Agence de notation", ""),
            ("Note court terme", ""),
            ("Note long terme", ""),
            ("Dernière date de notation", ""),
            ("Dividende versé régulièrement ?", "← Oui / Non"),
        ]),
        ("SOURCES D'INFORMATION", [
            ("Rapport annuel", "← Lien ou date"),
            ("Rapport trimestriel", ""),
            ("Site de la société", ""),
            ("Profil RichBourse", "← richbourse.com"),
            ("Profil BRVM", "← brvm.org"),
        ]),
    ]

    for section_title, fields in sections:
        style_section(ws, r, 1, section_title, merge_to_col=4)
        r += 1
        for label, placeholder in fields:
            style_label(ws, r, 1, label)
            cell = ws.cell(row=r, column=2, value=placeholder or None)
            if placeholder and placeholder.startswith("←"):
                cell.value = placeholder
                cell.font = ft(italic=True, color="808080", size=8)
            cell.fill = px(WHITE)
            cell.border = bd()
            cell.alignment = al("left", "center")
            for c in [3, 4]:
                ws.cell(row=r, column=c).border = bd()
            r += 1
        blank_row(ws, r); r += 1


# ══════════════════════════════════════════════════════════════════════════════
#  FEUILLE 3 : SYNTHESE
# ══════════════════════════════════════════════════════════════════════════════

def build_synthese(ws):
    set_col_widths(ws, {1: 28, 2: 16, 3: 20, 4: 16, 5: 20})

    r = 1
    style_title(ws, r, 1, "TABLEAU DE BORD — SYNTHÈSE DE L'ANALYSE",
                merge_to_col=5, size=13)
    r += 1
    style_title(ws, r, 1,
                "Société : [NOM]   |   Date : [JJ/MM/AAAA]   |   Prix analysé : [FCFA]",
                merge_to_col=5, bg=BLUE, size=10)
    r += 1
    blank_row(ws, r); r += 1

    # ── Scorecard fondamentale ───────────────────────────────────────────────
    style_section(ws, r, 1, "SCORECARD FONDAMENTALE", merge_to_col=5)
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5],
                 ["Critère", "Valeur", "Zone normale", "Signal", "Score (0–3)"])
    r += 1

    fundamental_criteria = [
        ("Croissance CA (N/N-4 %)", "← depuis feuille ETUDE", "> 10%"),
        ("Croissance RN (N/N-4 %)", "← depuis feuille ETUDE", "> 10%"),
        ("Marge nette moyenne (%)", "← depuis feuille ETUDE", "> 5%"),
        ("PER", "← depuis feuille ETUDE", "9 – 15"),
        ("PBR", "← depuis feuille ETUDE", "< 1 idéal"),
        ("ROE (%)", "← depuis feuille ETUDE", "> 10%"),
        ("Rendement dividende (%)", "← depuis feuille ETUDE", "> 3%"),
        ("RVC (PER × PBR)", "← depuis feuille ETUDE", "< 22"),
    ]

    for label, val_hint, zone in fundamental_criteria:
        style_label(ws, r, 1, label)
        cell = ws.cell(row=r, column=2, value=val_hint)
        cell.font = ft(italic=True, color="808080", size=8)
        cell.fill = px(WHITE); cell.border = bd()
        style_value(ws, r, 3, zone, bg=LORANGE)
        ws.cell(row=r, column=3).alignment = al("center", "center")
        cell_sig = ws.cell(row=r, column=4, value="← Saisir signal")
        cell_sig.font = ft(italic=True, color="808080", size=8)
        cell_sig.fill = px(WHITE); cell_sig.border = bd()
        cell_score = ws.cell(row=r, column=5, value=None)
        cell_score.fill = px(WHITE); cell_score.border = bd()
        cell_score.number_format = "0"
        r += 1

    # Score total fondamental
    SCORE_FOND_ROW = r
    style_label(ws, r, 1, "SCORE FONDAMENTAL TOTAL (/24)", bold=True)
    for c in [2, 3, 4]: ws.cell(row=r, column=c).border = bd()
    style_formula(ws, r, 5, f"=SUM(E{r-len(fundamental_criteria)}:E{r-1})",
                  fmt="0", bg=LGREEN, bold=True)
    r += 1
    blank_row(ws, r); r += 1

    # ── Scorecard technique ──────────────────────────────────────────────────
    style_section(ws, r, 1, "SCORECARD TECHNIQUE", merge_to_col=5)
    r += 1
    style_header(ws, r, [1, 2, 3, 4, 5],
                 ["Critère", "Signal observé", "Zone", "Interprétation", "Score"])
    r += 1

    technical_criteria = [
        ("MM20 + Bollinger + MACD + RSI + Volume",
         "← Score depuis ETUDE B1", "−5 à +5"),
        ("Prix Médians (PM 1 an + 2 ans)",
         "← Depuis ETUDE B2", "< PM = survente"),
        ("PER actualisé",
         "← Depuis ETUDE B3", "9–15 = normal"),
        ("PCD / Gordon-Shapiro",
         "← Depuis ETUDE B4/B5", "< PCD = survente"),
        ("Risk/Reward ratio",
         "← Depuis ETUDE B6", "≥ 2 = bon"),
    ]

    for label, val_hint, zone in technical_criteria:
        style_label(ws, r, 1, label)
        cell = ws.cell(row=r, column=2, value=val_hint)
        cell.font = ft(italic=True, color="808080", size=8)
        cell.fill = px(WHITE); cell.border = bd()
        style_value(ws, r, 3, zone, bg=LORANGE)
        ws.cell(row=r, column=3).alignment = al("center", "center")
        cell_int = ws.cell(row=r, column=4, value="← Saisir signal")
        cell_int.font = ft(italic=True, color="808080", size=8)
        cell_int.fill = px(WHITE); cell_int.border = bd()
        cell_sc = ws.cell(row=r, column=5, value=None)
        cell_sc.fill = px(WHITE); cell_sc.border = bd()
        r += 1

    SCORE_TECH_SYNTH_ROW = r
    style_label(ws, r, 1, "SCORE TECHNIQUE TOTAL (/15)", bold=True)
    for c in [2, 3, 4]: ws.cell(row=r, column=c).border = bd()
    style_formula(ws, r, 5, f"=SUM(E{r-len(technical_criteria)}:E{r-1})",
                  fmt="0", bg=LGREEN, bold=True)
    r += 1
    blank_row(ws, r); r += 1

    # ── Verdict global ──────────────────────────────────────────────────────
    style_section(ws, r, 1, "VERDICT GLOBAL", merge_to_col=5, bg=NAVY)
    r += 1

    style_header(ws, r, [1, 2, 3, 4, 5],
                 ["Composante", "Score", "Max", "% atteint", "Appréciation"])
    r += 1

    style_label(ws, r, 1, "Score Fondamental")
    style_formula(ws, r, 2, f"=E{SCORE_FOND_ROW}", fmt="0", bg=LGREEN, bold=True)
    style_value(ws, r, 3, 24, fmt="0", bg=GREY)
    style_formula(ws, r, 4, f"=IFERROR(E{SCORE_FOND_ROW}/24*100,\"\")",
                  fmt="0.0", bg=LGREEN)
    style_formula(ws, r, 5,
                  f'=IFERROR(IF(E{SCORE_FOND_ROW}/24>=0.7,"Solide",'
                  f'IF(E{SCORE_FOND_ROW}/24>=0.5,"Moyen","Faible")),"—")',
                  bg=WHITE)
    ws.cell(row=r, column=5).alignment = al("center", "center")
    r += 1

    style_label(ws, r, 1, "Score Technique")
    style_formula(ws, r, 2, f"=E{SCORE_TECH_SYNTH_ROW}", fmt="0", bg=LGREEN, bold=True)
    style_value(ws, r, 3, 15, fmt="0", bg=GREY)
    style_formula(ws, r, 4, f"=IFERROR(E{SCORE_TECH_SYNTH_ROW}/15*100,\"\")",
                  fmt="0.0", bg=LGREEN)
    style_formula(ws, r, 5,
                  f'=IFERROR(IF(E{SCORE_TECH_SYNTH_ROW}/15>=0.7,"Solide",'
                  f'IF(E{SCORE_TECH_SYNTH_ROW}/15>=0.5,"Moyen","Faible")),"—")',
                  bg=WHITE)
    ws.cell(row=r, column=5).alignment = al("center", "center")
    SCORE_TECH_ROW2 = r
    r += 1

    SCORE_TOTAL_ROW = r
    style_label(ws, r, 1, "SCORE TOTAL", bold=True)
    style_formula(ws, r, 2,
                  f"=IFERROR(E{SCORE_FOND_ROW}+E{SCORE_TECH_SYNTH_ROW},\"\")",
                  fmt="0", bg=LGREEN, bold=True)
    ws.cell(row=r, column=2).font = ft(bold=True, size=12, color=NAVY)
    style_value(ws, r, 3, 39, fmt="0", bg=GREY, bold=True)
    style_formula(ws, r, 4,
                  f"=IFERROR(B{r}/39*100,\"\")",
                  fmt="0.0", bg=LGREEN, bold=True)
    style_formula(ws, r, 5,
                  f'=IFERROR(IF(B{r}/39>=0.7,"ACHETER",'
                  f'IF(B{r}/39>=0.5,"SURVEILLER","EVITER")),"—")',
                  bg=WHITE, bold=True)
    ws.cell(row=r, column=5).font = ft(bold=True, size=11)
    ws.cell(row=r, column=5).alignment = al("center", "center")
    r += 1
    blank_row(ws, r); r += 1

    # ── Recommandation finale ────────────────────────────────────────────────
    style_section(ws, r, 1, "RECOMMANDATION FINALE (texte libre)", merge_to_col=5)
    r += 1
    cell = ws.cell(row=r, column=1)
    cell.value = ("← Résumer ici ta décision finale : "
                  "Acheter / Surveiller / Éviter, avec justification "
                  "(prix d'entrée, paliers, stop-loss, horizon).")
    cell.font = ft(italic=True, color="808080", size=9)
    cell.fill = px(WHITE)
    cell.alignment = al("left", "center", wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=5)
    ws.row_dimensions[r].height = 60


# ══════════════════════════════════════════════════════════════════════════════
#  FEUILLE 4 : FORMULE (référentiel)
# ══════════════════════════════════════════════════════════════════════════════

def build_formule(ws):
    set_col_widths(ws, {1: 28, 2: 32, 3: 20, 4: 40})

    r = 1
    style_title(ws, r, 1, "RÉFÉRENTIEL DES FORMULES & INDICATEURS BRVM",
                merge_to_col=4, size=13)
    r += 1
    blank_row(ws, r); r += 1

    formulas = [
        ("ANALYSE FONDAMENTALE", None, None, None),
        ("Croissance N/N-4",
         "(N − N-4) / |N-4| × 100",
         "%",
         "Comparer le CA ou RN sur 5 ans. Positif = croissance."),
        ("TCAM (CAGR)",
         "(N / N-4)^(1/4) − 1",
         "%",
         "Taux de Croissance Annuel Moyen sur 4 ans."),
        ("Marge Nette",
         "RN / CA × 100",
         "%",
         "< 2% = faible | 2–5% = moyen | > 5% = bon | > 10% = très bon"),
        ("PER",
         "Prix / BNPA",
         "fois",
         "Zone normale BRVM : 9–15. > 15 = surévalué, < 9 = sous-évalué"),
        ("VMC (Valeur Mathématique Comptable)",
         "Capitaux Propres / Nb titres",
         "FCFA",
         "Prix < VMC = action décotée (achat potentiellement intéressant)"),
        ("PBR (Price to Book)",
         "Capitalisation / Capitaux Propres",
         "fois",
         "< 1 = décote sur actif net | > 2 = prime significative"),
        ("ROE (Return on Equity)",
         "RN / Capitaux Propres × 100",
         "%",
         "< 5% = faible | 5–10% = moyen | 10–15% = bon | > 15% = excellent"),
        ("RVC",
         "PER × PBR",
         "",
         "< 22 = décote globale (Lynch). Indicateur de sous-valorisation composite."),
        ("Taux de distribution",
         "DVD / (BNPA × 0.85) × 100",
         "%",
         "0.85 = précompte 15% (BRVM). > 100% = distribution > bénéfice."),
        ("Rendement dividende (Yield)",
         "Dividende / Prix × 100",
         "%",
         "< 2% = faible | 2–5% = bon | > 5% = très attractif pour revenus"),
        ("ANALYSE TECHNIQUE", None, None, None),
        ("Prix Médian (PM)",
         "(Plus Haut + Plus Bas) / 2",
         "FCFA",
         "Prix < PM = zone survente. Calculer sur 1 an ET 2 ans séparément."),
        ("RSI",
         "Calculé sur graphique",
         "",
         "< 30 = survente | 30–70 = neutre | > 70 = surachat"),
        ("MACD",
         "Calculé sur graphique",
         "",
         "MACD > Signal = signal haussier (survente si récemment croisé à la hausse)"),
        ("Bandes de Bollinger",
         "Calculé sur graphique",
         "",
         "Prix sous bande inférieure = survente | Au-dessus bande sup = surachat"),
        ("PER actualisé",
         "Prix / BNPA × (1 + évol RN%)",
         "fois",
         "Intègre la dernière variation du RN pour estimer le PER forward"),
        ("PCD (Prix Cible Dividendes)",
         "100 × DVD_prév / Rendement_secteur%",
         "FCFA",
         "Rendement sectoriel BRVM : 5–8%. Ne s'applique que si dividende > 0."),
        ("Gordon-Shapiro",
         "P₀ = D₁ / (r − g)",
         "FCFA",
         "D₁ = dividende attendu | r = taux exigé | g = croissance dividendes"),
        ("GESTION DU RISQUE", None, None, None),
        ("Stop-Loss (10%)",
         "Prix d'achat × 0.90",
         "FCFA",
         "Limite de perte acceptable. Adapter selon profil de risque."),
        ("Risk/Reward",
         "Gain potentiel / Perte potentielle",
         "ratio",
         "< 1 = mauvais | 1–2 = acceptable | ≥ 2 = bon"),
        ("CMP (Coût Moyen Pondéré)",
         "Prix × 1.014",
         "FCFA",
         "Intègre la commission d'achat BRVM de 1.4%"),
        ("Taxe de cession",
         "Prix × 1.4%",
         "FCFA",
         "Taxe applicable à la vente sur BRVM."),
    ]

    style_header(ws, r, [1, 2, 3, 4],
                 ["Indicateur / Formule", "Expression mathématique",
                  "Unité", "Interprétation / Seuils"])
    r += 1

    for item in formulas:
        label, expr, unit, interp = item
        if expr is None:
            # C'est un titre de section
            style_section(ws, r, 1, label, merge_to_col=4, bg=BLUE)
            r += 1
            continue
        style_label(ws, r, 1, label, bold=True)
        style_value(ws, r, 2, expr, bg=LBLUE)
        ws.cell(row=r, column=2).alignment = al("center", "center")
        ws.cell(row=r, column=2).font = ft(size=9, color=NAVY)
        style_value(ws, r, 3, unit, bg=GREY)
        ws.cell(row=r, column=3).alignment = al("center", "center")
        cell_i = ws.cell(row=r, column=4, value=interp)
        cell_i.fill = px(WHITE)
        cell_i.font = ft(size=9, color=DARK)
        cell_i.alignment = al("left", "center", wrap=True)
        cell_i.border = bd()
        ws.row_dimensions[r].height = 28
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()

    # Renommer la feuille par défaut
    ws_etude = wb.active
    ws_etude.title = "ETUDE"

    ws_synthese = wb.create_sheet("SYNTHESE")
    ws_profil   = wb.create_sheet("PROFIL")
    ws_formule  = wb.create_sheet("FORMULE")

    build_etude(ws_etude)
    build_synthese(ws_synthese)
    build_profil(ws_profil)
    build_formule(ws_formule)

    # Figer les volets : première ligne de chaque feuille
    for ws in [ws_etude, ws_synthese, ws_profil, ws_formule]:
        ws.freeze_panes = "A2"

    # Sanitize ALL formula cells in the entire workbook (safety net)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell.value = sf(cell.value)

    out = "/home/user/Cloud-claude-tiaho/brvm-analyzer/etudes_actions_v2.xlsx"
    wb.save(out)
    print(f"Fichier cree : {out}")


if __name__ == "__main__":
    main()
