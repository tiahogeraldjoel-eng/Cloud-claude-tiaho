"""
Générateur du fichier Excel pré-rempli — portefeuille Tiaho au 09/06/2026
Source : état de portefeuille Coris Bourse
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from datetime import date

# ─── Données du portefeuille (extrait du PDF Coris Bourse 09/06/2026) ─────────
# (ticker, nom, qtite, cout_moyen, prix_marche_09062026, signal)
# Signal : H=fort gain/momentum, M=gain solide, L=faible/perte
PORTFOLIO = [
    ("SAFC",  "SAFCA CI",                 5,    3_745,  3_700,  "L"),
    ("STBC",  "SITAB CI",                23,   21_115, 21_900,  "L"),
    ("SOGC",  "SOGB CI",                 78,    8_020,  8_490,  "L"),
    ("SMBC",  "SMB CI",                  50,   11_817, 15_380,  "M"),
    ("NTLC",  "Nestlé CI",               20,   11_048, 14_225,  "M"),
    ("UNXC",  "Uniwax CI",              100,    1_957,  1_905,  "L"),
    ("SIVC",  "Air Liquide CI",          61,    2_757,  2_845,  "L"),
    ("BOAB",  "BOA Bénin",               40,    5_648,  8_745,  "H"),
    ("TTLC",  "TotalEnergies CI",       117,    2_834,  2_805,  "L"),
    ("BOAN",  "BOA Niger",               30,    2_677,  3_740,  "M"),
    ("ETIT",  "Ecobank TI (ETI)",      4000,       16,     33,  "H"),
    ("SDSC",  "Bolloré Transport CI",   250,    1_575,  2_000,  "M"),
    ("BOACI", "BOA CI",                 115,    7_488,  8_890,  "M"),
    ("BOABF", "BOA Burkina Faso",        67,    4_737,  5_595,  "M"),
    ("BOAS",  "BOA Sénégal",             20,    7_447,  7_400,  "L"),
    ("BOAM",  "BOA Mali",                30,    4_890,  4_630,  "L"),
    ("SIBC",  "SIB CI",                  78,    6_337,  8_510,  "H"),
    ("CBIBF", "Coris Bank BF",           30,   10_211, 21_500,  "H"),
    ("NSBC",  "NSIA Banque CI",          77,    8_301, 19_250,  "H"),
    ("ECOC",  "Ecobank CI",              32,   14_127, 16_800,  "M"),
    ("BICB",  "BIIC Bénin",              50,    5_182,  5_590,  "L"),
]

# ─── Styles ───────────────────────────────────────────────────────────────────

def bd():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_):   return PatternFill("solid", fgColor=hex_)
def ft(hex_="000000", bold=False, size=10):
    return Font(name="Calibri", color=hex_, bold=bold, size=size)

C_NAVY   = "1F3864"
C_BLUE2  = "2E75B6"
C_LIGHT  = "D9E1F2"
C_GOLD   = "FFF2CC"
C_GREEN  = "E2EFDA"
C_RED    = "FCE4D6"
C_ORANGE = "FCE4D6"
C_WHITE  = "FFFFFF"
C_YELLOW = "FFFACD"

def hdr(ws, r, c, val, bg=C_NAVY, fg="FFFFFF", bold=True, wrap=True, align="center"):
    cell = ws.cell(r, c, val)
    cell.fill   = fill(bg)
    cell.font   = Font(name="Calibri", color=fg, bold=bold, size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = bd()
    return cell

def dc(ws, r, c, val=None, bg=C_WHITE, bold=False, fmt=None, align="center"):
    cell = ws.cell(r, c, val)
    cell.fill   = fill(bg)
    cell.font   = ft(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = bd()
    if fmt: cell.number_format = fmt
    return cell

# ─── Feuille principale : Suivi portefeuille ─────────────────────────────────

def make_sheet(wb):
    ws = wb.create_sheet("📊 Portefeuille")
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:R1")
    t = ws["A1"]
    t.value = "BRVM — Suivi de Portefeuille  ·  Tiaho Gérald Joël  ·  Coris Bourse  ·  Réf. 09/06/2026"
    t.fill  = fill(C_NAVY)
    t.font  = Font(name="Calibri", color="FFFFFF", bold=True, size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:R2")
    s = ws["A2"]
    s.value = "⚡ Colonne K (Prix actuel) : mettre à jour chaque jour à 10h15 GMT après l'alerte Worker"
    s.fill  = fill(C_BLUE2)
    s.font  = Font(name="Calibri", color="FFFFFF", italic=True, size=9)
    s.alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes
    cols = [
        ("Date\nréf.",    9),   # A 1
        ("Ticker",        7),   # B 2
        ("Nom",          22),   # C 3
        ("Signal\n(H/M/L)", 8), # D 4
        ("Qtité",         7),   # E 5
        ("Coût\nMoyen",   9),   # F 6  — prix d'entrée historique
        ("Coût\nTotal",  11),   # G 7
        ("Stop\nOriginal\n−3%", 9),  # H 8  = F×0.97
        ("Stop\nProtection\n−5% actuel", 10), # I 9 = K×0.95
        ("Prochain\nObjectif\n+4% actuel", 10), # J 10 = K×1.04
        ("Prix\nActuel",  9),   # K 11  ← SAISIR ICI
        ("Valeur\nActuelle", 11), # L 12
        ("P&L\nFCFA",    11),   # M 13
        ("P&L\n%",        8),   # N 14
        ("Gain\ndepuis\norig.", 9), # O 15
        ("Statut",       20),   # P 16
        ("Valeur\nportefeuille\n% total", 9), # Q 17
        ("Notes",        18),   # R 18
    ]
    ws.row_dimensions[3].height = 36
    for i, (hd, w) in enumerate(cols, 1):
        hdr(ws, 3, i, hd, wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Lignes données ────────────────────────────────────────────────────────
    TOTAL_ROW = 3 + len(PORTFOLIO) + 1

    for idx, (ticker, nom, qtite, cout_moy, prix_mkt, sig) in enumerate(PORTFOLIO):
        r   = 4 + idx
        row = str(r)
        bg  = C_LIGHT if idx % 2 == 0 else C_WHITE

        pnl_pct = (prix_mkt - cout_moy) / cout_moy

        # Couleur signal
        sig_bg = {"H": "FFD7D7", "M": "FFE5CC", "L": "FFFACD"}.get(sig, C_WHITE)

        # A Date
        c = dc(ws, r, 1, date(2026, 6, 9), bg=bg)
        c.number_format = "DD/MM/YYYY"

        # B Ticker
        dc(ws, r, 2, ticker, bg=bg, bold=True, align="center")

        # C Nom
        dc(ws, r, 3, nom, bg=bg, align="left")

        # D Signal
        sig_label = {"H": "H 🔴", "M": "M 🟠", "L": "L 🟡"}.get(sig, sig)
        c = dc(ws, r, 4, sig_label, bg=sig_bg, bold=True, align="center")

        # E Qtité
        dc(ws, r, 5, qtite, bg=bg, fmt="#,##0", align="center")

        # F Coût Moyen (prix entrée historique)
        dc(ws, r, 6, cout_moy, bg=bg, fmt="#,##0 \"F\"", align="right")

        # G Coût Total = E × F
        dc(ws, r, 7,
           f"=E{row}*F{row}", bg=bg, fmt="#,##0", align="right")

        # H Stop Original = F × 0.97
        dc(ws, r, 8,
           f"=ROUND(F{row}*0.97,0)", bg="FFF3F3", fmt="#,##0 \"F\"", align="right")

        # I Stop Protection = K × 0.95  (protection des gains, trailing −5%)
        dc(ws, r, 9,
           f'=IF(K{row}=0,"",ROUND(K{row}*0.95,0))', bg="FFF3F3", fmt="#,##0 \"F\"", align="right")

        # J Prochain Objectif = K × 1.04
        dc(ws, r, 10,
           f'=IF(K{row}=0,"",ROUND(K{row}*1.04,0))', bg=C_GREEN, fmt="#,##0 \"F\"", align="right")

        # K Prix Actuel — pré-rempli + zone de saisie
        c = dc(ws, r, 11, prix_mkt, bg=C_YELLOW, fmt="#,##0", align="right", bold=True)
        c.fill = fill(C_YELLOW)

        # L Valeur Actuelle = K × E
        dc(ws, r, 12,
           f"=IF(K{row}=0,\"\",K{row}*E{row})", bg=bg, fmt="#,##0", align="right")

        # M P&L FCFA = L − G
        dc(ws, r, 13,
           f'=IF(K{row}=0,"",L{row}-G{row})', bg=bg, fmt="+#,##0;-#,##0;0", align="right", bold=True)

        # N P&L % = (K − F) / F
        dc(ws, r, 14,
           f'=IF(OR(F{row}=0,K{row}=0),"",(K{row}-F{row})/F{row})', bg=bg, fmt="+0.00%;-0.00%", align="center")

        # O Gain depuis origine = M en FCFA (redondant avec M mais utile pour tri)
        dc(ws, r, 15,
           f'=IF(M{row}="","",M{row})', bg=bg, fmt="+#,##0;-#,##0;0", align="right")

        # P Statut (basé sur P&L % depuis coût moyen)
        statut_formula = (
            f'=IF(OR(F{row}=0,K{row}=0),"⏳ En attente",'
            f'IF((K{row}-F{row})/F{row}<-0.05,"🛑 STOP FRANCHI — VENDRE",'
            f'IF((K{row}-F{row})/F{row}<-0.03,"⚠️ SURVEILLER — proche stop",'
            f'IF((K{row}-F{row})/F{row}<0.10,"⏳ TENIR",'
            f'IF((K{row}-F{row})/F{row}<0.30,"✅ BON GAIN — protéger",'
            f'IF((K{row}-F{row})/F{row}<0.80,"🚀 FORT GAIN — stop protection",'
            f'"💎 MULTIBAGGER — sécuriser impérativement"))))))'
        )
        dc(ws, r, 16, statut_formula, bg=bg, align="center")

        # Q % du total portefeuille
        dc(ws, r, 17,
           f'=IF(L{row}=0,"",L{row}/9316395)', bg=bg, fmt="0.00%", align="center")

        # R Notes
        dc(ws, r, 18, bg=bg, align="left")

    # ── Ligne TOTAL ───────────────────────────────────────────────────────────
    tr = str(TOTAL_ROW)
    ws.row_dimensions[TOTAL_ROW].height = 20
    ws.merge_cells(f"A{tr}:F{tr}")
    c = ws[f"A{tr}"]
    c.value = "TOTAL PORTEFEUILLE"
    c.fill  = fill(C_NAVY)
    c.font  = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = bd()

    # G — Coût global total
    c = dc(ws, TOTAL_ROW, 7,
           f"=SUM(G4:G{TOTAL_ROW-1})", bg=C_NAVY, bold=True, fmt="#,##0")
    c.font = Font(name="Calibri", color=C_GOLD, bold=True, size=10)

    # L — Valeur marchande totale
    c = dc(ws, TOTAL_ROW, 12,
           f"=SUM(L4:L{TOTAL_ROW-1})", bg=C_NAVY, bold=True, fmt="#,##0")
    c.font = Font(name="Calibri", color=C_GOLD, bold=True, size=10)

    # M — P&L total
    c = dc(ws, TOTAL_ROW, 13,
           f"=SUM(M4:M{TOTAL_ROW-1})", bg=C_NAVY, bold=True, fmt="+#,##0;-#,##0")
    c.font = Font(name="Calibri", color="00FF88", bold=True, size=11)

    # N — P&L % global
    c = dc(ws, TOTAL_ROW, 14,
           f"=IF(G{tr}=0,\"\",(L{tr}-G{tr})/G{tr})", bg=C_NAVY, bold=True, fmt="+0.00%;-0.00%")
    c.font = Font(name="Calibri", color="00FF88", bold=True, size=11)

    # ── Mise en forme conditionnelle ──────────────────────────────────────────
    n_rows = len(PORTFOLIO)
    data_range_p = f"P4:P{3+n_rows}"
    data_range_m = f"M4:M{3+n_rows}"

    ws.conditional_formatting.add(data_range_p, FormulaRule(
        formula=['ISNUMBER(SEARCH("STOP",P4))'],
        fill=fill("FFCCCC"), font=Font(bold=True, color="C00000", name="Calibri", size=9)))
    ws.conditional_formatting.add(data_range_p, FormulaRule(
        formula=['ISNUMBER(SEARCH("SURVEILLER",P4))'],
        fill=fill("FFE0B2"), font=Font(bold=True, color="E65100", name="Calibri", size=9)))
    ws.conditional_formatting.add(data_range_p, FormulaRule(
        formula=['ISNUMBER(SEARCH("BON GAIN",P4))'],
        fill=fill("CCFFCC"), font=Font(bold=True, color="1B5E20", name="Calibri", size=9)))
    ws.conditional_formatting.add(data_range_p, FormulaRule(
        formula=['ISNUMBER(SEARCH("FORT",P4))'],
        fill=fill("B3E5FC"), font=Font(bold=True, color="0D47A1", name="Calibri", size=9)))
    ws.conditional_formatting.add(data_range_p, FormulaRule(
        formula=['ISNUMBER(SEARCH("MULTI",P4))'],
        fill=fill("FFD700"), font=Font(bold=True, color="7B3F00", name="Calibri", size=9)))

    ws.conditional_formatting.add(data_range_m, FormulaRule(
        formula=["M4>0"], fill=fill("CCFFCC"),
        font=Font(bold=True, color="1B5E20", name="Calibri", size=10)))
    ws.conditional_formatting.add(data_range_m, FormulaRule(
        formula=["M4<0"], fill=fill("FFCCCC"),
        font=Font(bold=True, color="C00000", name="Calibri", size=10)))

    ws.freeze_panes = "A4"

    # ── Légende en bas ────────────────────────────────────────────────────────
    leg_start = TOTAL_ROW + 2
    ws.merge_cells(f"A{leg_start}:R{leg_start}")
    c = ws[f"A{leg_start}"]
    c.value = "LÉGENDE STATUTS"
    c.fill  = fill(C_NAVY)
    c.font  = Font(name="Calibri", color="FFFFFF", bold=True, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = bd()

    legends = [
        (leg_start+1, "⏳ TENIR",
         "Gain <10% — position normale, attendre la prochaine alerte Worker", C_WHITE, "000000"),
        (leg_start+2, "✅ BON GAIN — protéger",
         "Gain +10% à +30% — placer un stop de protection à −5% du cours actuel (col I)", "CCFFCC", "1B5E20"),
        (leg_start+3, "🚀 FORT GAIN — stop protection",
         "Gain +30% à +80% — OBLIGATOIRE : actualiser le stop protection chaque semaine", "B3E5FC", "0D47A1"),
        (leg_start+4, "💎 MULTIBAGGER — sécuriser",
         "Gain >80% (ETIT, CBIBF, NSBC) — envisager une prise de bénéfices partielle", "FFD700", "7B3F00"),
        (leg_start+5, "⚠️ SURVEILLER — proche stop",
         "Perte −3% à −5% — surveiller quotidiennement", "FFE0B2", "E65100"),
        (leg_start+6, "🛑 STOP FRANCHI — VENDRE",
         "Perte >−5% du coût moyen — VENDRE immédiatement pour limiter les pertes", "FFCCCC", "C00000"),
    ]
    for r_l, stat, expl, bg_l, fg_l in legends:
        ws.merge_cells(f"A{r_l}:D{r_l}")
        c = ws[f"A{r_l}"]
        c.value = stat
        c.fill  = fill(bg_l)
        c.font  = Font(name="Calibri", bold=True, color=fg_l, size=9)
        c.alignment = Alignment(vertical="center")
        c.border = bd()
        ws.merge_cells(f"E{r_l}:R{r_l}")
        c2 = ws[f"E{r_l}"]
        c2.value = expl
        c2.fill  = fill(bg_l)
        c2.font  = Font(name="Calibri", color=fg_l, size=9)
        c2.alignment = Alignment(vertical="center", horizontal="left")
        c2.border = bd()
        ws.row_dimensions[r_l].height = 16

    return ws

# ─── Feuille 2 : Analyse des positions ───────────────────────────────────────

def make_analyse(wb):
    ws = wb.create_sheet("📈 Analyse", 1)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "Analyse du Portefeuille au 09/06/2026 — Classement par P&L"
    t.fill  = fill(C_NAVY)
    t.font  = Font(name="Calibri", color="FFFFFF", bold=True, size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Calculer P&L pour chaque position
    data = []
    for ticker, nom, qtite, cout_moy, prix_mkt, sig in PORTFOLIO:
        pnl_fcfa = (prix_mkt - cout_moy) * qtite
        pnl_pct  = (prix_mkt - cout_moy) / cout_moy * 100
        val_mkt  = prix_mkt * qtite
        stop_prot = round(prix_mkt * 0.95)
        next_obj  = round(prix_mkt * 1.04)
        data.append((pnl_pct, ticker, nom, qtite, cout_moy, prix_mkt, pnl_fcfa, val_mkt, stop_prot, next_obj, sig))

    # Trier par P&L% décroissant
    data.sort(key=lambda x: x[0], reverse=True)

    # En-têtes
    headers = [
        ("Rang", 5), ("Ticker", 8), ("Nom", 22), ("Signal", 8),
        ("P&L %", 9), ("P&L FCFA", 12), ("Coût Moy", 10),
        ("Prix Actuel", 10), ("Stop Protect.", 11), ("Prochain Obj.", 11),
        ("Recommandation", 28),
    ]
    ws.row_dimensions[2].height = 22
    for i, (h, w) in enumerate(headers, 1):
        hdr(ws, 2, i, h, bg="2E4057")
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, (pnl_pct, ticker, nom, qtite, cout_moy, prix_mkt, pnl_fcfa, val_mkt, stop_prot, next_obj, sig) in enumerate(data):
        r   = 3 + idx
        row = str(r)
        bg  = C_LIGHT if idx % 2 == 0 else C_WHITE

        if pnl_pct > 80:
            reco = "💎 Prise partielle recommandée — gain exceptionnel à sécuriser"
            bg_r = "FFF9C4"
        elif pnl_pct > 30:
            reco = "🚀 Actualiser le stop protection (−5% du cours actuel)"
            bg_r = "E3F2FD"
        elif pnl_pct > 10:
            reco = "✅ Tenir — placer stop protection à " + f"{stop_prot:,} F"
            bg_r = "F1F8E9"
        elif pnl_pct > 0:
            reco = "⏳ Tenir — surveiller lors de l'alerte quotidienne"
            bg_r = C_WHITE
        elif pnl_pct > -3:
            reco = "⚠️ Proche du seuil de stop — surveiller de près"
            bg_r = "FFF3E0"
        else:
            reco = "🛑 Stop franchi — décision de vente à envisager"
            bg_r = "FFEBEE"

        def cell(c, v, fmt=None, bold=False, align="center", bg2=bg):
            cc = ws.cell(r, c, v)
            cc.fill   = fill(bg2)
            cc.font   = ft(bold=bold)
            cc.alignment = Alignment(horizontal=align, vertical="center")
            cc.border = bd()
            if fmt: cc.number_format = fmt
            return cc

        cell(1, idx+1, align="center")
        cell(2, ticker, bold=True)
        cell(3, nom, align="left")
        sig_label = {"H":"H 🔴","M":"M 🟠","L":"L 🟡"}.get(sig, sig)
        cell(4, sig_label,
             bg2={"H":"FFD7D7","M":"FFE5CC","L":"FFFACD"}.get(sig, bg))
        cell(5, pnl_pct/100, fmt="+0.00%;-0.00%", bold=True,
             bg2="CCFFCC" if pnl_pct>0 else "FFCCCC")
        cell(6, pnl_fcfa,    fmt="+#,##0;-#,##0", bold=True,
             bg2="CCFFCC" if pnl_fcfa>0 else "FFCCCC")
        cell(7, cout_moy,    fmt="#,##0 \"F\"")
        cell(8, prix_mkt,    fmt="#,##0 \"F\"", bold=True)
        cell(9, stop_prot,   fmt="#,##0 \"F\"", bg2="FFF3F3")
        cell(10, next_obj,   fmt="#,##0 \"F\"", bg2="F1F8E9")
        cell(11, reco, align="left", bg2=bg_r)

    # Total P&L
    total_pnl  = sum([(pm - cm) * q for _, _, q, cm, pm, _ in PORTFOLIO])
    total_cost = sum([cm * q         for _, _, q, cm, pm, _ in PORTFOLIO])
    total_val  = sum([pm * q         for _, _, q, cm, pm, _ in PORTFOLIO])
    total_pct  = (total_val - total_cost) / total_cost * 100

    tr = 3 + len(PORTFOLIO) + 1
    ws.merge_cells(f"A{tr}:D{tr}")
    c = ws[f"A{tr}"]
    c.value = f"TOTAL PORTEFEUILLE — Coût : {total_cost:,.0f} F  ·  Valeur : {total_val:,.0f} F"
    c.fill  = fill(C_NAVY)
    c.font  = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = bd()

    c5 = ws.cell(tr, 5, total_pct/100)
    c5.fill = fill(C_NAVY)
    c5.font = Font(name="Calibri", color="00FF88", bold=True, size=11)
    c5.number_format = "+0.00%;-0.00%"
    c5.alignment = Alignment(horizontal="center")
    c5.border = bd()

    c6 = ws.cell(tr, 6, total_pnl)
    c6.fill = fill(C_NAVY)
    c6.font = Font(name="Calibri", color="00FF88", bold=True, size=11)
    c6.number_format = "+#,##0;-#,##0"
    c6.alignment = Alignment(horizontal="center")
    c6.border = bd()

    ws.freeze_panes = "A3"
    return ws

# ─── Main ─────────────────────────────────────────────────────────────────────

wb = Workbook()
del wb[wb.sheetnames[0]]

make_sheet(wb)
make_analyse(wb)

out = "BRVM_Portefeuille_Tiaho_09062026.xlsx"
wb.save(out)

# Résumé console
print(f"✅  {out}")
print()
total_pnl  = sum([(pm-cm)*q  for _,_,q,cm,pm,_ in PORTFOLIO])
total_cost = sum([cm*q        for _,_,q,cm,pm,_ in PORTFOLIO])
total_val  = sum([pm*q        for _,_,q,cm,pm,_ in PORTFOLIO])
print(f"Portfolio  : {len(PORTFOLIO)} lignes")
print(f"Coût total : {total_cost:>12,.0f} FCFA")
print(f"Valeur mkt : {total_val:>12,.0f} FCFA")
print(f"P&L total  : {total_pnl:>+12,.0f} FCFA  ({total_pnl/total_cost*100:+.1f}%)")
print()

# Alerte positions à risque
risk = [(t,n,q,cm,pm) for t,n,q,cm,pm,s in PORTFOLIO if (pm-cm)/cm < -0.03]
if risk:
    print("⚠️  POSITIONS STOP FRANCHI :")
    for t,n,q,cm,pm in risk:
        print(f"   {t:6s} {n:25s}  {(pm-cm)/cm*100:+.1f}%  P&L: {(pm-cm)*q:+,.0f} F")
